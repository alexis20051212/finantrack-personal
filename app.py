from flask import Flask, render_template, request, redirect, url_for, session, flash, make_response, jsonify
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, date, timedelta
from functools import wraps
import mysql.connector
from mysql.connector import Error
from calendar import monthrange
from io import BytesIO
from exchange_api import exchange_api, COMMON_CURRENCIES
from fred_api import fred_api, FRED_SERIES
from db_config import get_db_connection, get_dict_cursor
import os

app = Flask(__name__)
app.secret_key = 'FinanTrack_MySQL_Clave_Segura_2024'

# ==================== DECORADOR ====================
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Por favor inicia sesión', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# ==================== FUNCIONES DE UTILIDAD ====================

def crear_notificacion(usuario_id, mensaje, tipo='info'):
    """Crea una notificación para un usuario"""
    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO notificaciones (usuario_id, mensaje, tipo)
                VALUES (%s, %s, %s)
            """, (usuario_id, mensaje, tipo))
            conn.commit()
            cursor.close()
            conn.close()
        except Exception as e:
            print(f"Error creando notificación: {e}")
            if conn:
                conn.close()

def verificar_presupuestos_notificacion(usuario_id):
    """Verifica si algún presupuesto fue excedido"""
    conn = get_db_connection()
    if not conn:
        return
    
    try:
        cursor = get_dict_cursor(conn)
        mes_actual = datetime.now().month
        anio_actual = datetime.now().year
        
        if os.environ.get('RENDER'):
            cursor.execute("""
                SELECT c.nombre as categoria, p.limite, COALESCE(SUM(m.monto), 0) as gastado
                FROM presupuestos p
                JOIN categorias c ON p.categoria_id = c.id
                LEFT JOIN movimientos m ON c.id = m.categoria_id 
                    AND m.tipo = 'gasto' AND m.usuario_id = %s
                    AND EXTRACT(MONTH FROM m.fecha) = %s AND EXTRACT(YEAR FROM m.fecha) = %s
                WHERE p.usuario_id = %s AND p.mes = %s AND p.anio = %s
                GROUP BY c.id, c.nombre, p.limite
                HAVING COALESCE(SUM(m.monto), 0) > p.limite
            """, (usuario_id, mes_actual, anio_actual, usuario_id, mes_actual, anio_actual))
        else:
            cursor.execute("""
                SELECT c.nombre as categoria, p.limite, COALESCE(SUM(m.monto), 0) as gastado
                FROM presupuestos p
                JOIN categorias c ON p.categoria_id = c.id
                LEFT JOIN movimientos m ON c.id = m.categoria_id 
                    AND m.tipo = 'gasto' AND m.usuario_id = %s
                    AND MONTH(m.fecha) = %s AND YEAR(m.fecha) = %s
                WHERE p.usuario_id = %s AND p.mes = %s AND p.anio = %s
                GROUP BY c.id, c.nombre, p.limite
                HAVING gastado > limite
            """, (usuario_id, mes_actual, anio_actual, usuario_id, mes_actual, anio_actual))
        
        presupuestos_excedidos = cursor.fetchall()
        cursor.close()
        conn.close()
        
        for p in presupuestos_excedidos:
            porcentaje = (p['gastado'] / p['limite'] * 100)
            mensaje = f"Has excedido el presupuesto de {p['categoria']} en {porcentaje:.0f}% (${p['gastado'] - p['limite']:,.2f})"
            crear_notificacion(usuario_id, mensaje, 'peligro')
    except Exception as e:
        print(f"Error verificando presupuestos: {e}")
        if conn:
            conn.close()

def verificar_recordatorios_pendientes():
    """Verifica y notifica recordatorios próximos"""
    conn = get_db_connection()
    if not conn:
        return
    
    try:
        cursor = get_dict_cursor(conn)
        fecha_limite = (datetime.now() + timedelta(days=3)).strftime('%Y-%m-%d')
        
        if os.environ.get('RENDER'):
            cursor.execute("""
                SELECT r.*, u.email, u.nombre as usuario_nombre
                FROM recordatorios r
                JOIN users u ON r.usuario_id = u.id
                WHERE r.fecha <= %s AND r.fecha >= CURRENT_DATE AND r.completado = FALSE
            """, (fecha_limite,))
        else:
            cursor.execute("""
                SELECT r.*, u.email, u.nombre as usuario_nombre
                FROM recordatorios r
                JOIN users u ON r.usuario_id = u.id
                WHERE r.fecha <= %s AND r.fecha >= CURDATE() AND r.completado = FALSE
            """, (fecha_limite,))
        
        recordatorios = cursor.fetchall()
        cursor.close()
        conn.close()
        
        for r in recordatorios:
            dias = (r['fecha'] - datetime.now().date()).days
            if dias == 0:
                mensaje = f"HOY: {r['titulo']}"
            elif dias == 1:
                mensaje = f"Mañana: {r['titulo']}"
            else:
                mensaje = f"En {dias} días: {r['titulo']}"
            
            if r['descripcion']:
                mensaje += f" - {r['descripcion']}"
            
            crear_notificacion(r['usuario_id'], mensaje, 'alerta')
    except Exception as e:
        print(f"Error verificando recordatorios: {e}")
        if conn:
            conn.close()

# ==================== RUTAS ====================

@app.route('/')
def home():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/cambiar-tema', methods=['GET', 'POST'])
def cambiar_tema():
    if request.method == 'POST':
        data = request.get_json()
        nuevo_tema = data.get('tema', 'dark')
        session['tema'] = nuevo_tema
        return {'tema': nuevo_tema}
    else:
        tema_actual = session.get('tema', 'dark')
        return {'tema': tema_actual}

# ==================== REGISTRO Y LOGIN ====================

@app.route('/register', methods=['GET', 'POST'])
def register():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        nombre = request.form.get('nombre', '').strip()
        email = request.form.get('email', '').strip().lower()
        telefono = request.form.get('telefono', '').strip()
        password = request.form.get('password', '')
        confirm_password = request.form.get('confirm_password', '')
        default_currency = request.form.get('default_currency', 'USD')
        
        errores = []
        
        if not nombre or not email or not password:
            errores.append('Todos los campos son obligatorios')
        
        if telefono and not telefono.isdigit():
            errores.append('El teléfono solo debe contener números')
        
        if len(password) < 8:
            errores.append('La contraseña debe tener al menos 8 caracteres')
        
        if not any(c.isupper() for c in password):
            errores.append('La contraseña debe contener al menos una letra mayúscula')
        
        caracteres_especiales = "!@#$%^&*(),.?\":{}|<>"
        if not any(c in caracteres_especiales for c in password):
            errores.append('La contraseña debe contener al menos un carácter especial (!@#$%^&*)')
        
        if password != confirm_password:
            errores.append('Las contraseñas no coinciden')
        
        if errores:
            for error in errores:
                flash(error, 'danger')
            return render_template('register.html', currencies=COMMON_CURRENCIES)
        
        conn = get_db_connection()
        if not conn:
            flash('Error de conexión a la base de datos', 'danger')
            return render_template('register.html', currencies=COMMON_CURRENCIES)
        
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT id FROM users WHERE email = %s", (email,))
            if cursor.fetchone():
                flash('Este correo ya está registrado', 'danger')
                cursor.close()
                conn.close()
                return render_template('register.html', currencies=COMMON_CURRENCIES)
            
            hashed_password = generate_password_hash(password)
            cursor.execute("""
                INSERT INTO users (nombre, email, telefono, password, default_currency)
                VALUES (%s, %s, %s, %s, %s)
            """, (nombre, email, telefono if telefono else None, hashed_password, default_currency))
            conn.commit()
            
            if os.environ.get('RENDER'):
                cursor.execute("SELECT lastval()")
                user_id = cursor.fetchone()[0]
            else:
                user_id = cursor.lastrowid
            
            cursor.close()
            conn.close()
            
            crear_notificacion(user_id, '¡Bienvenido a FinanTrack! Comienza registrando tus primeros movimientos.', 'exito')
            flash('¡Registro exitoso!', 'success')
            return redirect(url_for('login'))
            
        except Exception as e:
            print(f"Error en registro: {e}")
            flash('Error en la base de datos', 'danger')
            if conn:
                conn.close()
            return render_template('register.html', currencies=COMMON_CURRENCIES)
    
    return render_template('register.html', currencies=COMMON_CURRENCIES)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '')
        
        if not email or not password:
            flash('Email y contraseña son obligatorios', 'danger')
            return render_template('login.html')
        
        conn = get_db_connection()
        if not conn:
            flash('Error de conexión', 'danger')
            return render_template('login.html')
        
        try:
            cursor = get_dict_cursor(conn)
            cursor.execute("SELECT id, nombre, password, default_currency FROM users WHERE email = %s", (email,))
            user = cursor.fetchone()
            cursor.close()
            conn.close()
            
            if user and check_password_hash(user['password'], password):
                session['user_id'] = user['id']
                session['user_name'] = user['nombre']
                session['default_currency'] = user['default_currency'] or 'USD'
                flash(f'¡Bienvenido {user["nombre"]}!', 'success')
                return redirect(url_for('dashboard'))
            else:
                flash('Email o contraseña incorrectos', 'danger')
        except Exception as e:
            print(f"Error en login: {e}")
            flash('Error al iniciar sesión', 'danger')
            if conn:
                conn.close()
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('Sesión cerrada', 'info')
    return redirect(url_for('login'))

# ==================== DASHBOARD ====================

@app.route('/dashboard')
@login_required
def dashboard():
    try:
        conn = get_db_connection()
        if not conn:
            flash('Error de conexión', 'danger')
            return redirect(url_for('login'))
        
        cursor = get_dict_cursor(conn)
        user_id = session['user_id']
        user_currency = session.get('default_currency', 'USD')
        
        mes_actual = datetime.now().month
        anio_actual = datetime.now().year
        mes_anterior = mes_actual - 1 if mes_actual > 1 else 12
        anio_anterior = anio_actual if mes_actual > 1 else anio_actual - 1
        
        try:
            verificar_presupuestos_notificacion(user_id)
        except:
            pass
        try:
            verificar_recordatorios_pendientes()
        except:
            pass
        
        # Datos del mes actual
        if os.environ.get('RENDER'):
            cursor.execute("""
                SELECT 
                    COALESCE(SUM(CASE WHEN tipo = 'ingreso' THEN monto ELSE 0 END), 0) as ingresos,
                    COALESCE(SUM(CASE WHEN tipo = 'gasto' THEN monto ELSE 0 END), 0) as gastos
                FROM movimientos
                WHERE usuario_id = %s 
                AND EXTRACT(MONTH FROM fecha) = %s 
                AND EXTRACT(YEAR FROM fecha) = %s
            """, (user_id, mes_actual, anio_actual))
        else:
            cursor.execute("""
                SELECT 
                    COALESCE(SUM(CASE WHEN tipo = 'ingreso' THEN monto ELSE 0 END), 0) as ingresos,
                    COALESCE(SUM(CASE WHEN tipo = 'gasto' THEN monto ELSE 0 END), 0) as gastos
                FROM movimientos
                WHERE usuario_id = %s 
                AND MONTH(fecha) = %s 
                AND YEAR(fecha) = %s
            """, (user_id, mes_actual, anio_actual))
        
        row = cursor.fetchone()
        total_ingresos = float(row['ingresos'] or 0)
        total_gastos = float(row['gastos'] or 0)
        balance = total_ingresos - total_gastos
        
        # Datos del mes anterior
        if os.environ.get('RENDER'):
            cursor.execute("""
                SELECT 
                    COALESCE(SUM(CASE WHEN tipo = 'ingreso' THEN monto ELSE 0 END), 0) as ingresos,
                    COALESCE(SUM(CASE WHEN tipo = 'gasto' THEN monto ELSE 0 END), 0) as gastos
                FROM movimientos
                WHERE usuario_id = %s 
                AND EXTRACT(MONTH FROM fecha) = %s 
                AND EXTRACT(YEAR FROM fecha) = %s
            """, (user_id, mes_anterior, anio_anterior))
        else:
            cursor.execute("""
                SELECT 
                    COALESCE(SUM(CASE WHEN tipo = 'ingreso' THEN monto ELSE 0 END), 0) as ingresos,
                    COALESCE(SUM(CASE WHEN tipo = 'gasto' THEN monto ELSE 0 END), 0) as gastos
                FROM movimientos
                WHERE usuario_id = %s 
                AND MONTH(fecha) = %s 
                AND YEAR(fecha) = %s
            """, (user_id, mes_anterior, anio_anterior))
        
        row = cursor.fetchone()
        ingresos_anterior = float(row['ingresos'] or 0)
        gastos_anterior = float(row['gastos'] or 0)
        
        variacion_ingresos = ((total_ingresos - ingresos_anterior) / ingresos_anterior * 100) if ingresos_anterior > 0 else 0
        variacion_gastos = ((total_gastos - gastos_anterior) / gastos_anterior * 100) if gastos_anterior > 0 else 0
        
        # Total transacciones
        cursor.execute("SELECT COUNT(*) as total FROM movimientos WHERE usuario_id = %s", (user_id,))
        total_transacciones = cursor.fetchone()['total'] or 0
        
        # Gastos diario promedio
        dias_del_mes = monthrange(anio_actual, mes_actual)[1]
        gasto_diario_promedio = total_gastos / dias_del_mes if dias_del_mes > 0 else 0
        
        # Porcentaje de meta de ahorro
        porcentaje_meta = 100 if total_ingresos > 0 and balance >= 0 else (balance / total_ingresos * 100) if total_ingresos > 0 else 0
        porcentaje_meta = max(0, min(100, porcentaje_meta))
        
        # Datos para gráficos (últimos 6 meses)
        meses_nombres = []
        ingresos_mensuales = []
        gastos_mensuales = []
        
        for i in range(5, -1, -1):
            mes_num = mes_actual - i
            año_num = anio_actual
            if mes_num <= 0:
                mes_num += 12
                año_num -= 1
            
            meses_nombres.append(date(2000, mes_num, 1).strftime('%b'))
            
            if os.environ.get('RENDER'):
                cursor.execute("""
                    SELECT COALESCE(SUM(CASE WHEN tipo = 'ingreso' THEN monto ELSE 0 END), 0) as ingresos,
                           COALESCE(SUM(CASE WHEN tipo = 'gasto' THEN monto ELSE 0 END), 0) as gastos
                    FROM movimientos
                    WHERE usuario_id = %s 
                    AND EXTRACT(MONTH FROM fecha) = %s 
                    AND EXTRACT(YEAR FROM fecha) = %s
                """, (user_id, mes_num, año_num))
            else:
                cursor.execute("""
                    SELECT COALESCE(SUM(CASE WHEN tipo = 'ingreso' THEN monto ELSE 0 END), 0) as ingresos,
                           COALESCE(SUM(CASE WHEN tipo = 'gasto' THEN monto ELSE 0 END), 0) as gastos
                    FROM movimientos
                    WHERE usuario_id = %s 
                    AND MONTH(fecha) = %s 
                    AND YEAR(fecha) = %s
                """, (user_id, mes_num, año_num))
            
            row = cursor.fetchone()
            ingresos_mensuales.append(float(row['ingresos'] or 0))
            gastos_mensuales.append(float(row['gastos'] or 0))
        
        # Top categorías - CORREGIDO PARA POSTGRESQL
        if os.environ.get('RENDER'):
            cursor.execute("""
                SELECT c.nombre as categoria_nombre, COALESCE(SUM(m.monto), 0) as total
                FROM categorias c
                JOIN movimientos m ON c.id = m.categoria_id 
                    AND m.tipo = 'gasto' 
                    AND m.usuario_id = %s
                    AND EXTRACT(MONTH FROM m.fecha) = %s 
                    AND EXTRACT(YEAR FROM m.fecha) = %s
                GROUP BY c.id, c.nombre
                HAVING COALESCE(SUM(m.monto), 0) > 0
                ORDER BY total DESC
                LIMIT 6
            """, (user_id, mes_actual, anio_actual))
        else:
            cursor.execute("""
                SELECT c.nombre as categoria_nombre, COALESCE(SUM(m.monto), 0) as total
                FROM categorias c
                JOIN movimientos m ON c.id = m.categoria_id 
                    AND m.tipo = 'gasto' 
                    AND m.usuario_id = %s
                    AND MONTH(m.fecha) = %s 
                    AND YEAR(m.fecha) = %s
                GROUP BY c.id, c.nombre
                HAVING total > 0
                ORDER BY total DESC
                LIMIT 6
            """, (user_id, mes_actual, anio_actual))
        
        categorias_top = cursor.fetchall()
        categorias_nombres = [cat['categoria_nombre'] for cat in categorias_top]
        categorias_totales = [float(cat['total']) for cat in categorias_top]
        
        # Gastos por categoría - CORREGIDO PARA POSTGRESQL
        if os.environ.get('RENDER'):
            cursor.execute("""
                SELECT c.nombre as categoria_nombre, COALESCE(SUM(m.monto), 0) as total
                FROM categorias c
                LEFT JOIN movimientos m ON c.id = m.categoria_id 
                    AND m.tipo = 'gasto' 
                    AND m.usuario_id = %s
                    AND EXTRACT(MONTH FROM m.fecha) = %s 
                    AND EXTRACT(YEAR FROM m.fecha) = %s
                GROUP BY c.id, c.nombre
                HAVING COALESCE(SUM(m.monto), 0) > 0
                ORDER BY total DESC
            """, (user_id, mes_actual, anio_actual))
        else:
            cursor.execute("""
                SELECT c.nombre as categoria_nombre, COALESCE(SUM(m.monto), 0) as total
                FROM categorias c
                LEFT JOIN movimientos m ON c.id = m.categoria_id 
                    AND m.tipo = 'gasto' 
                    AND m.usuario_id = %s
                    AND MONTH(m.fecha) = %s 
                    AND YEAR(m.fecha) = %s
                GROUP BY c.id, c.nombre
                HAVING total > 0
                ORDER BY total DESC
            """, (user_id, mes_actual, anio_actual))
        
        gastos_por_categoria_raw = cursor.fetchall()
        gastos_por_categoria = []
        if total_gastos > 0:
            for item in gastos_por_categoria_raw:
                gastos_por_categoria.append({
                    'categoria_nombre': item['categoria_nombre'],
                    'total': float(item['total']),
                    'porcentaje': (float(item['total']) / total_gastos) * 100
                })
        
        # Ingresos por categoría - CORREGIDO PARA POSTGRESQL
        if os.environ.get('RENDER'):
            cursor.execute("""
                SELECT c.nombre as categoria_nombre, COALESCE(SUM(m.monto), 0) as total
                FROM categorias c
                LEFT JOIN movimientos m ON c.id = m.categoria_id 
                    AND m.tipo = 'ingreso' 
                    AND m.usuario_id = %s
                    AND EXTRACT(MONTH FROM m.fecha) = %s 
                    AND EXTRACT(YEAR FROM m.fecha) = %s
                GROUP BY c.id, c.nombre
                HAVING COALESCE(SUM(m.monto), 0) > 0
                ORDER BY total DESC
            """, (user_id, mes_actual, anio_actual))
        else:
            cursor.execute("""
                SELECT c.nombre as categoria_nombre, COALESCE(SUM(m.monto), 0) as total
                FROM categorias c
                LEFT JOIN movimientos m ON c.id = m.categoria_id 
                    AND m.tipo = 'ingreso' 
                    AND m.usuario_id = %s
                    AND MONTH(m.fecha) = %s 
                    AND YEAR(m.fecha) = %s
                GROUP BY c.id, c.nombre
                HAVING total > 0
                ORDER BY total DESC
            """, (user_id, mes_actual, anio_actual))
        
        ingresos_por_categoria_raw = cursor.fetchall()
        ingresos_por_categoria = []
        if total_ingresos > 0:
            for item in ingresos_por_categoria_raw:
                ingresos_por_categoria.append({
                    'categoria_nombre': item['categoria_nombre'],
                    'total': float(item['total']),
                    'porcentaje': (float(item['total']) / total_ingresos) * 100
                })
        
        # Top 3 categorías para tarjetas
        top_3 = gastos_por_categoria[:3] if gastos_por_categoria else []
        
        # Últimos movimientos
        if os.environ.get('RENDER'):
            cursor.execute("""
                SELECT m.id, m.tipo, m.monto, m.descripcion, 
                       TO_CHAR(m.fecha, 'DD/MM/YYYY') as fecha, 
                       c.nombre as categoria_nombre, m.currency
                FROM movimientos m
                JOIN categorias c ON m.categoria_id = c.id
                WHERE m.usuario_id = %s
                ORDER BY m.fecha DESC, m.id DESC
                LIMIT 10
            """, (user_id,))
        else:
            cursor.execute("""
                SELECT m.id, m.tipo, m.monto, m.descripcion, 
                       DATE_FORMAT(m.fecha, '%d/%m/%Y') as fecha, 
                       c.nombre as categoria_nombre, m.currency
                FROM movimientos m
                JOIN categorias c ON m.categoria_id = c.id
                WHERE m.usuario_id = %s
                ORDER BY m.fecha DESC, m.id DESC
                LIMIT 10
            """, (user_id,))
        
        ultimos_movimientos = cursor.fetchall()
        
        # Notificaciones no leídas
        cursor.execute("""
            SELECT COUNT(*) as total FROM notificaciones 
            WHERE usuario_id = %s AND leido = FALSE
        """, (user_id,))
        notificaciones_no_leidas = cursor.fetchone()['total'] or 0
        
        cursor.close()
        conn.close()
        
        # Consejo del zorro
        consejo_zorro = {}
        if total_ingresos == 0 and total_gastos == 0:
            consejo_zorro = {
                'icono': '🦊',
                'titulo': '¡Bienvenido a FinanTrack!',
                'mensaje': 'Comienza registrando tus primeros ingresos y gastos.',
                'color': '#60A5FA',
                'accion': 'agregar_movimiento',
                'texto_boton': 'Primer Movimiento'
            }
        elif balance > 500:
            consejo_zorro = {
                'icono': '🏆',
                'titulo': '¡Excelente ahorro!',
                'mensaje': f'Has ahorrado ${balance:.2f} este mes. ¡Sigue así!',
                'color': '#34D399',
                'accion': None,
                'texto_boton': None
            }
        elif balance > 0:
            consejo_zorro = {
                'icono': '👍',
                'titulo': 'Buen trabajo',
                'mensaje': f'Has ahorrado ${balance:.2f} este mes.',
                'color': '#FBBF24',
                'accion': None,
                'texto_boton': None
            }
        elif balance == 0 and total_ingresos > 0:
            consejo_zorro = {
                'icono': '⚖️',
                'titulo': 'Balance en cero',
                'mensaje': 'Tus ingresos igualan a tus gastos. Busca formas de ahorrar.',
                'color': '#FBBF24',
                'accion': None,
                'texto_boton': None
            }
        elif balance < 0:
            consejo_zorro = {
                'icono': '⚠️',
                'titulo': 'Cuidado',
                'mensaje': f'Tus gastos superan tus ingresos por ${abs(balance):.2f}.',
                'color': '#F87171',
                'accion': None,
                'texto_boton': None
            }
        elif total_gastos == 0 and total_ingresos > 0:
            consejo_zorro = {
                'icono': '💰',
                'titulo': '¡Puro ingreso!',
                'mensaje': 'Registra tus gastos para tener una visión completa.',
                'color': '#34D399',
                'accion': 'agregar_movimiento',
                'texto_boton': 'Registrar Gasto'
            }
        elif total_ingresos == 0 and total_gastos > 0:
            consejo_zorro = {
                'icono': '📉',
                'titulo': 'Sin ingresos',
                'mensaje': 'Agrega tus fuentes de ingreso.',
                'color': '#FBBF24',
                'accion': 'agregar_movimiento',
                'texto_boton': 'Registrar Ingreso'
            }
        else:
            consejo_zorro = {
                'icono': '🦊',
                'titulo': 'Consejo del día',
                'mensaje': 'Mantén un registro constante de tus gastos.',
                'color': '#60A5FA',
                'accion': None,
                'texto_boton': None
            }
        
        return render_template('dashboard.html', 
                             nombre=session['user_name'],
                             total_ingresos=total_ingresos,
                             total_gastos=total_gastos,
                             balance=balance,
                             ahorro=balance if balance > 0 else 0,
                             variacion_ingresos=variacion_ingresos,
                             variacion_gastos=variacion_gastos,
                             gasto_diario_promedio=gasto_diario_promedio,
                             total_transacciones=total_transacciones,
                             porcentaje_meta=porcentaje_meta,
                             top_categorias=top_3,
                             gastos_por_categoria=gastos_por_categoria,
                             ingresos_por_categoria=ingresos_por_categoria,
                             movimientos=ultimos_movimientos,
                             meses_nombres=meses_nombres,
                             ingresos_mensuales=ingresos_mensuales,
                             gastos_mensuales=gastos_mensuales,
                             categorias_nombres=categorias_nombres,
                             categorias_totales=categorias_totales,
                             consejo_zorro=consejo_zorro,
                             notificaciones_no_leidas=notificaciones_no_leidas,
                             user_currency=user_currency,
                             currencies=COMMON_CURRENCIES)
    
    except Exception as e:
        print(f"Error en dashboard: {e}")
        import traceback
        traceback.print_exc()
        flash('Error al cargar el dashboard', 'danger')
        return redirect(url_for('logout'))

# ==================== CRUD MOVIMIENTOS ====================

@app.route('/agregar-movimiento', methods=['GET', 'POST'])
@login_required
def agregar_movimiento():
    conn = get_db_connection()
    if not conn:
        flash('Error de conexión', 'danger')
        return redirect(url_for('dashboard'))
    
    try:
        cursor = get_dict_cursor(conn)
        cursor.execute("SELECT id, nombre FROM categorias ORDER BY nombre")
        categorias = cursor.fetchall()
        cursor.close()
        conn.close()
    except Exception as e:
        print(f"Error obteniendo categorías: {e}")
        if conn:
            conn.close()
        flash('Error al cargar categorías', 'danger')
        return redirect(url_for('dashboard'))
    
    user_currency = session.get('default_currency', 'USD')
    
    if request.method == 'POST':
        tipo = request.form.get('tipo')
        categoria_id = request.form.get('categoria_id')
        monto = float(request.form.get('monto'))
        descripcion = request.form.get('descripcion', '')
        fecha = request.form.get('fecha')
        currency = request.form.get('currency', user_currency)
        
        if not fecha:
            fecha = datetime.now().strftime('%Y-%m-%d')
        
        conn = get_db_connection()
        if not conn:
            flash('Error de conexión', 'danger')
            return redirect(url_for('dashboard'))
        
        try:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO movimientos (usuario_id, tipo, monto, categoria_id, descripcion, fecha, currency)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (session['user_id'], tipo, monto, categoria_id, descripcion, fecha, currency))
            conn.commit()
            cursor.close()
            conn.close()
            
            crear_notificacion(session['user_id'], f"Movimiento registrado: {tipo} de ${monto:,.2f} {currency}", 'exito')
            flash('Movimiento agregado correctamente', 'success')
            return redirect(url_for('dashboard'))
        except Exception as e:
            print(f"Error agregando movimiento: {e}")
            if conn:
                conn.close()
            flash('Error al agregar movimiento', 'danger')
    
    return render_template('movimiento.html', categorias=categorias, movimiento=None, 
                         currencies=COMMON_CURRENCIES, user_currency=user_currency)

@app.route('/editar-movimiento/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_movimiento(id):
    conn = get_db_connection()
    if not conn:
        flash('Error de conexión', 'danger')
        return redirect(url_for('dashboard'))
    
    user_currency = session.get('default_currency', 'USD')
    
    if request.method == 'POST':
        tipo = request.form.get('tipo')
        categoria_id = request.form.get('categoria_id')
        monto = float(request.form.get('monto'))
        descripcion = request.form.get('descripcion', '')
        fecha = request.form.get('fecha')
        currency = request.form.get('currency', user_currency)
        
        try:
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE movimientos 
                SET tipo=%s, monto=%s, categoria_id=%s, descripcion=%s, fecha=%s, currency=%s
                WHERE id=%s AND usuario_id=%s
            """, (tipo, monto, categoria_id, descripcion, fecha, currency, id, session['user_id']))
            conn.commit()
            cursor.close()
            conn.close()
            
            crear_notificacion(session['user_id'], f"Movimiento editado correctamente", 'info')
            flash('Movimiento actualizado', 'success')
            return redirect(url_for('dashboard'))
        except Exception as e:
            print(f"Error editando movimiento: {e}")
            if conn:
                conn.close()
            flash('Error al editar movimiento', 'danger')
    
    try:
        cursor = get_dict_cursor(conn)
        cursor.execute("SELECT * FROM movimientos WHERE id=%s AND usuario_id=%s", (id, session['user_id']))
        movimiento = cursor.fetchone()
        cursor.execute("SELECT id, nombre FROM categorias ORDER BY nombre")
        categorias = cursor.fetchall()
        cursor.close()
        conn.close()
    except Exception as e:
        print(f"Error obteniendo datos: {e}")
        if conn:
            conn.close()
        flash('Error al cargar datos', 'danger')
        return redirect(url_for('dashboard'))
    
    if not movimiento:
        flash('Movimiento no encontrado', 'danger')
        return redirect(url_for('dashboard'))
    
    return render_template('movimiento.html', categorias=categorias, movimiento=movimiento, 
                         currencies=COMMON_CURRENCIES, user_currency=user_currency)

@app.route('/eliminar-movimiento/<int:id>')
@login_required
def eliminar_movimiento(id):
    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM movimientos WHERE id=%s AND usuario_id=%s", (id, session['user_id']))
            conn.commit()
            cursor.close()
            conn.close()
            
            crear_notificacion(session['user_id'], f"Movimiento eliminado", 'info')
            flash('Movimiento eliminado', 'success')
        except Exception as e:
            print(f"Error eliminando movimiento: {e}")
            if conn:
                conn.close()
            flash('Error al eliminar movimiento', 'danger')
    else:
        flash('Error de conexión', 'danger')
    
    return redirect(url_for('dashboard'))

# ==================== METAS DE AHORRO ====================

@app.route('/metas')
@login_required
def listar_metas():
    conn = get_db_connection()
    if not conn:
        flash('Error de conexión', 'danger')
        return redirect(url_for('dashboard'))
    
    try:
        cursor = get_dict_cursor(conn)
        user_currency = session.get('default_currency', 'USD')
        
        # Obtener todas las metas sin cálculos complejos en SQL
        cursor.execute("""
            SELECT id, usuario_id, nombre, monto_objetivo, monto_actual, 
                   fecha_limite, currency, fecha_creacion
            FROM metas 
            WHERE usuario_id = %s
            ORDER BY id DESC
        """, (session['user_id'],))
        
        metas = cursor.fetchall()
        cursor.close()
        conn.close()
        
        # Procesar cada meta en Python (compatible con ambos motores)
        for meta in metas:
            # Calcular porcentaje
            if meta['monto_objetivo'] and meta['monto_objetivo'] > 0:
                meta['porcentaje'] = round((meta['monto_actual'] / meta['monto_objetivo'] * 100), 2)
            else:
                meta['porcentaje'] = 0
            
            # Calcular días restantes
            if meta['fecha_limite']:
                hoy = datetime.now().date()
                dias = (meta['fecha_limite'] - hoy).days
                meta['dias_restantes'] = dias if dias >= 0 else 0
            else:
                meta['dias_restantes'] = None
        
        return render_template('metas.html', metas=metas, user_currency=user_currency, currencies=COMMON_CURRENCIES)
    except Exception as e:
        print(f"Error listando metas: {e}")
        import traceback
        traceback.print_exc()
        if conn:
            conn.close()
        flash('Error al cargar metas', 'danger')
        return redirect(url_for('dashboard'))

@app.route('/metas/nueva', methods=['POST'])
@login_required
def nueva_meta():
    nombre = request.form.get('nombre')
    monto_objetivo = float(request.form.get('monto_objetivo'))
    fecha_limite = request.form.get('fecha_limite')
    currency = request.form.get('currency', session.get('default_currency', 'USD'))
    
    conn = get_db_connection()
    if not conn:
        flash('Error de conexión', 'danger')
        return redirect(url_for('listar_metas'))
    
    try:
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO metas (usuario_id, nombre, monto_objetivo, fecha_limite, currency)
            VALUES (%s, %s, %s, %s, %s)
        """, (session['user_id'], nombre, monto_objetivo, fecha_limite if fecha_limite else None, currency))
        conn.commit()
        cursor.close()
        conn.close()
        
        crear_notificacion(session['user_id'], f"¡Nueva meta creada: {nombre} por ${monto_objetivo:,.2f} {currency}!", 'exito')
        flash('Meta de ahorro creada exitosamente', 'success')
    except Exception as e:
        print(f"Error creando meta: {e}")
        if conn:
            conn.close()
        flash('Error al crear meta', 'danger')
    
    return redirect(url_for('listar_metas'))

@app.route('/metas/aportar/<int:id>', methods=['POST'])
@login_required
def aportar_meta(id):
    try:
        # Obtener y validar el monto
        monto_str = request.form.get('monto', '0')
        monto = float(monto_str)
        descripcion = request.form.get('descripcion', '')
        
        if monto <= 0:
            flash('El monto debe ser mayor a 0', 'danger')
            return redirect(url_for('listar_metas'))
        
        conn = get_db_connection()
        if not conn:
            flash('Error de conexión a la base de datos', 'danger')
            return redirect(url_for('listar_metas'))
        
        cursor = conn.cursor()
        
        # Verificar la meta
        cursor.execute("""
            SELECT id, monto_objetivo, monto_actual 
            FROM metas 
            WHERE id = %s AND usuario_id = %s
        """, (id, session['user_id']))
        
        meta = cursor.fetchone()
        
        if not meta:
            cursor.close()
            conn.close()
            flash('Meta no encontrada', 'danger')
            return redirect(url_for('listar_metas'))
        
        # Extraer y convertir valores
        meta_id = int(meta[0])
        monto_objetivo = float(str(meta[1])) if meta[1] is not None else 0.0
        monto_actual = float(str(meta[2])) if meta[2] is not None else 0.0
        
        # Calcular nuevo monto
        nuevo_monto = monto_actual + monto
        
        # Actualizar la meta
        cursor.execute("""
            UPDATE metas 
            SET monto_actual = %s
            WHERE id = %s AND usuario_id = %s
        """, (nuevo_monto, meta_id, session['user_id']))
        
        # Insertar aportación
        fecha_actual = datetime.now().strftime('%Y-%m-%d')
        cursor.execute("""
            INSERT INTO aportaciones_meta (meta_id, monto, fecha, descripcion)
            VALUES (%s, %s, %s, %s)
        """, (meta_id, monto, fecha_actual, descripcion))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        # Notificar
        if nuevo_monto >= monto_objetivo:
            crear_notificacion(
                session['user_id'], 
                f"🎉 ¡Meta completada! Ahorraste ${nuevo_monto:,.2f} de ${monto_objetivo:,.2f}", 
                'exito'
            )
            flash('🎉 ¡Meta completada! Felicitaciones.', 'success')
        else:
            porcentaje = (nuevo_monto / monto_objetivo * 100) if monto_objetivo > 0 else 0
            crear_notificacion(
                session['user_id'], 
                f"Aportaste ${monto:,.2f}. Progreso: {porcentaje:.1f}%", 
                'info'
            )
            flash(f'¡Aportación de ${monto:,.2f} registrada!', 'success')
        
        return redirect(url_for('listar_metas'))
        
    except ValueError as e:
        print(f"Error de valor: {e}")
        flash('El monto ingresado no es válido', 'danger')
        return redirect(url_for('listar_metas'))
    except Exception as e:
        print(f"Error registrando aportación: {e}")
        import traceback
        traceback.print_exc()
        flash('Error al registrar la aportación', 'danger')
        if 'conn' in locals() and conn:
            conn.close()
        return redirect(url_for('listar_metas'))

@app.route('/metas/eliminar/<int:id>')
@login_required
def eliminar_meta(id):
    conn = get_db_connection()
    if not conn:
        flash('Error de conexión', 'danger')
        return redirect(url_for('listar_metas'))
    
    try:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM metas WHERE id = %s AND usuario_id = %s", (id, session['user_id']))
        conn.commit()
        cursor.close()
        conn.close()
        
        crear_notificacion(session['user_id'], f"Meta eliminada", 'info')
        flash('Meta eliminada', 'success')
    except Exception as e:
        print(f"Error eliminando meta: {e}")
        if conn:
            conn.close()
        flash('Error al eliminar meta', 'danger')
    
    return redirect(url_for('listar_metas'))

# ==================== PRESUPUESTOS ====================

@app.route('/presupuestos')
@login_required
def listar_presupuestos():
    conn = get_db_connection()
    if not conn:
        flash('Error de conexión', 'danger')
        return redirect(url_for('dashboard'))
    
    try:
        cursor = get_dict_cursor(conn)
        user_currency = session.get('default_currency', 'USD')
        
        mes_actual = datetime.now().month
        anio_actual = datetime.now().year
        
        cursor.execute("""
            SELECT p.*, c.nombre as categoria_nombre
            FROM presupuestos p
            JOIN categorias c ON p.categoria_id = c.id
            WHERE p.usuario_id = %s AND p.mes = %s AND p.anio = %s
        """, (session['user_id'], mes_actual, anio_actual))
        presupuestos = cursor.fetchall()
        
        if os.environ.get('RENDER'):
            cursor.execute("""
                SELECT c.id as categoria_id, c.nombre, COALESCE(SUM(m.monto), 0) as gastado
                FROM categorias c
                LEFT JOIN movimientos m ON c.id = m.categoria_id 
                    AND m.tipo = 'gasto' 
                    AND m.usuario_id = %s
                    AND EXTRACT(MONTH FROM m.fecha) = %s 
                    AND EXTRACT(YEAR FROM m.fecha) = %s
                GROUP BY c.id, c.nombre
            """, (session['user_id'], mes_actual, anio_actual))
        else:
            cursor.execute("""
                SELECT c.id as categoria_id, c.nombre, COALESCE(SUM(m.monto), 0) as gastado
                FROM categorias c
                LEFT JOIN movimientos m ON c.id = m.categoria_id 
                    AND m.tipo = 'gasto' 
                    AND m.usuario_id = %s
                    AND MONTH(m.fecha) = %s 
                    AND YEAR(m.fecha) = %s
                GROUP BY c.id, c.nombre
            """, (session['user_id'], mes_actual, anio_actual))
        
        gastos_reales = cursor.fetchall()
        
        for p in presupuestos:
            for g in gastos_reales:
                if p['categoria_id'] == g['categoria_id']:
                    p['gastado'] = g['gastado']
                    p['porcentaje'] = (g['gastado'] / p['limite'] * 100) if p['limite'] > 0 else 0
                    break
            else:
                p['gastado'] = 0
                p['porcentaje'] = 0
        
        categorias_sin_presupuesto = []
        for g in gastos_reales:
            if not any(p['categoria_id'] == g['categoria_id'] for p in presupuestos):
                categorias_sin_presupuesto.append(g)
        
        cursor.close()
        conn.close()
        
        return render_template('presupuestos.html', 
                             presupuestos=presupuestos,
                             categorias_sin_presupuesto=categorias_sin_presupuesto,
                             mes=mes_actual,
                             anio=anio_actual,
                             user_currency=user_currency,
                             currencies=COMMON_CURRENCIES)
    except Exception as e:
        print(f"Error listando presupuestos: {e}")
        if conn:
            conn.close()
        flash('Error al cargar presupuestos', 'danger')
        return redirect(url_for('dashboard'))

@app.route('/presupuestos/nuevo', methods=['POST'])
@login_required
def nuevo_presupuesto():
    categoria_id = request.form.get('categoria_id')
    limite = float(request.form.get('limite'))
    mes = int(request.form.get('mes', datetime.now().month))
    anio = int(request.form.get('anio', datetime.now().year))
    currency = request.form.get('currency', session.get('default_currency', 'USD'))
    
    conn = get_db_connection()
    if not conn:
        flash('Error de conexión', 'danger')
        return redirect(url_for('listar_presupuestos'))
    
    try:
        cursor = conn.cursor()
        
        if os.environ.get('RENDER'):
            cursor.execute("""
                INSERT INTO presupuestos (usuario_id, categoria_id, mes, anio, limite, currency)
                VALUES (%s, %s, %s, %s, %s, %s)
                ON CONFLICT (usuario_id, categoria_id, mes, anio) 
                DO UPDATE SET limite = %s, currency = %s
            """, (session['user_id'], categoria_id, mes, anio, limite, currency, limite, currency))
        else:
            cursor.execute("""
                INSERT INTO presupuestos (usuario_id, categoria_id, mes, anio, limite, currency)
                VALUES (%s, %s, %s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE limite = %s, currency = %s
            """, (session['user_id'], categoria_id, mes, anio, limite, currency, limite, currency))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        crear_notificacion(session['user_id'], f"Presupuesto creado con límite de ${limite:,.2f} {currency}", 'exito')
        flash('Presupuesto guardado', 'success')
    except Exception as e:
        print(f"Error creando presupuesto: {e}")
        if conn:
            conn.close()
        flash('Error al guardar presupuesto', 'danger')
    
    return redirect(url_for('listar_presupuestos'))

@app.route('/presupuestos/eliminar/<int:id>')
@login_required
def eliminar_presupuesto(id):
    conn = get_db_connection()
    if not conn:
        flash('Error de conexión', 'danger')
        return redirect(url_for('listar_presupuestos'))
    
    try:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM presupuestos WHERE id = %s AND usuario_id = %s", (id, session['user_id']))
        conn.commit()
        cursor.close()
        conn.close()
        
        crear_notificacion(session['user_id'], f"Presupuesto eliminado", 'info')
        flash('Presupuesto eliminado', 'success')
    except Exception as e:
        print(f"Error eliminando presupuesto: {e}")
        if conn:
            conn.close()
        flash('Error al eliminar presupuesto', 'danger')
    
    return redirect(url_for('listar_presupuestos'))

# ==================== RECORDATORIOS ====================

@app.route('/recordatorios')
@login_required
def listar_recordatorios():
    conn = get_db_connection()
    if not conn:
        flash('Error de conexión', 'danger')
        return redirect(url_for('dashboard'))
    
    try:
        cursor = get_dict_cursor(conn)
        
        cursor.execute("""
            SELECT * FROM recordatorios 
            WHERE usuario_id = %s AND completado = FALSE
            ORDER BY fecha ASC
        """, (session['user_id'],))
        recordatorios_pendientes = cursor.fetchall()
        
        cursor.execute("""
            SELECT * FROM recordatorios 
            WHERE usuario_id = %s AND completado = TRUE
            ORDER BY fecha DESC
            LIMIT 10
        """, (session['user_id'],))
        recordatorios_completados = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return render_template('recordatorios.html', 
                             pendientes=recordatorios_pendientes,
                             completados=recordatorios_completados,
                             now_date=datetime.now().date(),
                             timedelta=timedelta)
    except Exception as e:
        print(f"Error listando recordatorios: {e}")
        if conn:
            conn.close()
        flash('Error al cargar recordatorios', 'danger')
        return redirect(url_for('dashboard'))

@app.route('/recordatorios/nuevo', methods=['POST'])
@login_required
def nuevo_recordatorio():
    titulo = request.form.get('titulo')
    descripcion = request.form.get('descripcion', '')
    fecha = request.form.get('fecha')
    tipo = request.form.get('tipo', 'recordatorio')
    
    if not fecha:
        flash('La fecha es obligatoria', 'danger')
        return redirect(url_for('listar_recordatorios'))
    
    conn = get_db_connection()
    if not conn:
        flash('Error de conexión', 'danger')
        return redirect(url_for('listar_recordatorios'))
    
    try:
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO recordatorios (usuario_id, titulo, descripcion, fecha, tipo)
            VALUES (%s, %s, %s, %s, %s)
        """, (session['user_id'], titulo, descripcion, fecha, tipo))
        conn.commit()
        cursor.close()
        conn.close()
        
        crear_notificacion(session['user_id'], f"Recordatorio creado: {titulo}", 'exito')
        flash('Recordatorio creado correctamente', 'success')
    except Exception as e:
        print(f"Error creando recordatorio: {e}")
        if conn:
            conn.close()
        flash('Error al crear recordatorio', 'danger')
    
    return redirect(url_for('listar_recordatorios'))

@app.route('/recordatorios/completar/<int:id>')
@login_required
def completar_recordatorio(id):
    conn = get_db_connection()
    if not conn:
        flash('Error de conexión', 'danger')
        return redirect(url_for('listar_recordatorios'))
    
    try:
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE recordatorios 
            SET completado = TRUE 
            WHERE id = %s AND usuario_id = %s
        """, (id, session['user_id']))
        conn.commit()
        cursor.close()
        conn.close()
        
        crear_notificacion(session['user_id'], f"¡Completaste un recordatorio!", 'exito')
        flash('Recordatorio marcado como completado', 'success')
    except Exception as e:
        print(f"Error completando recordatorio: {e}")
        if conn:
            conn.close()
        flash('Error al completar recordatorio', 'danger')
    
    return redirect(url_for('listar_recordatorios'))

@app.route('/recordatorios/eliminar/<int:id>')
@login_required
def eliminar_recordatorio(id):
    conn = get_db_connection()
    if not conn:
        flash('Error de conexión', 'danger')
        return redirect(url_for('listar_recordatorios'))
    
    try:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM recordatorios WHERE id = %s AND usuario_id = %s", (id, session['user_id']))
        conn.commit()
        cursor.close()
        conn.close()
        
        crear_notificacion(session['user_id'], f"Recordatorio eliminado", 'info')
        flash('Recordatorio eliminado', 'success')
    except Exception as e:
        print(f"Error eliminando recordatorio: {e}")
        if conn:
            conn.close()
        flash('Error al eliminar recordatorio', 'danger')
    
    return redirect(url_for('listar_recordatorios'))

# ==================== NOTIFICACIONES ====================

@app.route('/notificaciones')
@login_required
def listar_notificaciones():
    conn = get_db_connection()
    if not conn:
        flash('Error de conexión', 'danger')
        return redirect(url_for('dashboard'))
    
    try:
        cursor = get_dict_cursor(conn)
        cursor.execute("""
            SELECT * FROM notificaciones 
            WHERE usuario_id = %s 
            ORDER BY fecha_creacion DESC
        """, (session['user_id'],))
        notificaciones = cursor.fetchall()
        cursor.close()
        conn.close()
        
        return render_template('notificaciones.html', notificaciones=notificaciones)
    except Exception as e:
        print(f"Error listando notificaciones: {e}")
        if conn:
            conn.close()
        flash('Error al cargar notificaciones', 'danger')
        return redirect(url_for('dashboard'))

@app.route('/notificaciones/marcar-leida/<int:id>')
@login_required
def marcar_notificacion_leida(id):
    conn = get_db_connection()
    if not conn:
        flash('Error de conexión', 'danger')
        return redirect(url_for('listar_notificaciones'))
    
    try:
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE notificaciones 
            SET leido = TRUE 
            WHERE id = %s AND usuario_id = %s
        """, (id, session['user_id']))
        conn.commit()
        cursor.close()
        conn.close()
    except Exception as e:
        print(f"Error marcando notificación: {e}")
        if conn:
            conn.close()
    
    return redirect(url_for('listar_notificaciones'))

@app.route('/notificaciones/marcar-todas')
@login_required
def marcar_todas_notificaciones():
    conn = get_db_connection()
    if not conn:
        flash('Error de conexión', 'danger')
        return redirect(url_for('listar_notificaciones'))
    
    try:
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE notificaciones 
            SET leido = TRUE 
            WHERE usuario_id = %s
        """, (session['user_id'],))
        conn.commit()
        cursor.close()
        conn.close()
        
        flash('Todas las notificaciones marcadas como leídas', 'success')
    except Exception as e:
        print(f"Error marcando todas las notificaciones: {e}")
        if conn:
            conn.close()
        flash('Error al marcar notificaciones', 'danger')
    
    return redirect(url_for('listar_notificaciones'))

@app.route('/api/notificaciones/no-leidas')
@login_required
def api_notificaciones_no_leidas():
    conn = get_db_connection()
    if not conn:
        return {'total': 0}
    
    try:
        cursor = get_dict_cursor(conn)
        cursor.execute("""
            SELECT COUNT(*) as total FROM notificaciones 
            WHERE usuario_id = %s AND leido = FALSE
        """, (session['user_id'],))
        resultado = cursor.fetchone()
        cursor.close()
        conn.close()
        
        return {'total': resultado['total'] if resultado else 0}
    except Exception as e:
        print(f"Error obteniendo notificaciones no leídas: {e}")
        if conn:
            conn.close()
        return {'total': 0}

# ==================== EXCHANGE RATE API ====================

@app.route('/cambiar-divisa', methods=['POST'])
@login_required
def cambiar_divisa():
    data = request.get_json()
    currency = data.get('currency', 'USD')
    
    if currency:
        session['default_currency'] = currency
        
        conn = get_db_connection()
        if conn:
            try:
                cursor = conn.cursor()
                cursor.execute("UPDATE users SET default_currency = %s WHERE id = %s", (currency, session['user_id']))
                conn.commit()
                cursor.close()
                conn.close()
            except Exception as e:
                print(f"Error cambiando divisa: {e}")
                if conn:
                    conn.close()
        
        return {'success': True, 'currency': currency}
    return {'success': False, 'error': 'Moneda no válida'}, 400

@app.route('/api/exchange-rates')
@login_required
def api_exchange_rates():
    base = request.args.get('base', 'USD')
    rates_data = exchange_api.get_exchange_rates(base)
    
    if rates_data:
        return {
            'success': True,
            'base': base,
            'rates': rates_data.get('conversion_rates', {}),
            'last_updated': rates_data.get('time_last_update_utc', '')
        }
    return {'success': False, 'error': 'No se pudieron obtener las tasas de cambio'}

@app.route('/api/convert-currency', methods=['POST'])
@login_required
def api_convert_currency():
    data = request.get_json()
    amount = data.get('amount')
    from_currency = data.get('from_currency', 'USD')
    to_currency = data.get('to_currency', 'MXN')
    
    if not amount:
        return {'success': False, 'error': 'Monto requerido'}
    
    converted = exchange_api.convert_currency(amount, from_currency, to_currency)
    
    if converted is not None:
        return {
            'success': True,
            'amount': amount,
            'from_currency': from_currency,
            'to_currency': to_currency,
            'converted_amount': converted,
            'rate': converted / amount if amount > 0 else 0
        }
    return {'success': False, 'error': 'Error al convertir la moneda'}

@app.route('/api/currencies')
@login_required
def api_currencies():
    return {
        'success': True,
        'currencies': [{'code': code, 'name': name} for code, name in COMMON_CURRENCIES]
    }

@app.route('/configuracion-divisas')
@login_required
def configuracion_divisas():
    user_currency = session.get('default_currency', 'USD')
    return render_template('divisas.html', 
                         currencies=COMMON_CURRENCIES,
                         user_currency=user_currency)

# ==================== FRED API ====================

@app.route('/fred')
@login_required
def fred_dashboard():
    return render_template('fred.html', 
                         series=FRED_SERIES,
                         user_currency=session.get('default_currency', 'USD'))

@app.route('/api/fred/series')
@login_required
def api_fred_series():
    return {
        'success': True,
        'series': [{'id': k, 'name': v} for k, v in FRED_SERIES.items()]
    }

@app.route('/api/fred/series/<series_id>/info')
@login_required
def api_fred_series_info(series_id):
    info = fred_api.get_series_info(series_id)
    if info:
        return {'success': True, 'info': info}
    return {'success': False, 'error': 'Serie no encontrada'}

@app.route('/api/fred/series/<series_id>/data')
@login_required
def api_fred_series_data(series_id):
    days = request.args.get('days', 365, type=int)
    data = fred_api.get_historical_data(series_id, days)
    
    if data:
        return {
            'success': True,
            'series_id': series_id,
            'data': data,
            'latest_value': data[0]['value'] if data else None
        }
    return {'success': False, 'error': 'No se pudieron obtener los datos'}

@app.route('/api/fred/series/<series_id>/latest')
@login_required
def api_fred_series_latest(series_id):
    value = fred_api.get_latest_value(series_id)
    if value is not None:
        return {'success': True, 'series_id': series_id, 'latest_value': value}
    return {'success': False, 'error': 'No se pudo obtener el valor'}

@app.route('/api/fred/search')
@login_required
def api_fred_search():
    query = request.args.get('q', '')
    limit = request.args.get('limit', 10, type=int)
    
    if not query:
        return {'success': False, 'error': 'Se requiere un término de búsqueda'}
    
    results = fred_api.search_series(query, limit)
    return {'success': True, 'results': results}

# ==================== GRÁFICOS ====================

@app.route('/api/datos-graficos')
@login_required
def datos_graficos():
    conn = get_db_connection()
    if not conn:
        return {'error': 'Error de conexión'}
    
    try:
        cursor = get_dict_cursor(conn)
        anio_actual = datetime.now().year
        user_id = session['user_id']
        
        meses = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
        ingresos_por_mes = []
        gastos_por_mes = []
        
        for mes in range(1, 13):
            if os.environ.get('RENDER'):
                cursor.execute("""
                    SELECT COALESCE(SUM(CASE WHEN tipo = 'ingreso' THEN monto ELSE 0 END), 0) as ingresos,
                           COALESCE(SUM(CASE WHEN tipo = 'gasto' THEN monto ELSE 0 END), 0) as gastos
                    FROM movimientos
                    WHERE usuario_id = %s AND EXTRACT(MONTH FROM fecha) = %s AND EXTRACT(YEAR FROM fecha) = %s
                """, (user_id, mes, anio_actual))
            else:
                cursor.execute("""
                    SELECT COALESCE(SUM(CASE WHEN tipo = 'ingreso' THEN monto ELSE 0 END), 0) as ingresos,
                           COALESCE(SUM(CASE WHEN tipo = 'gasto' THEN monto ELSE 0 END), 0) as gastos
                    FROM movimientos
                    WHERE usuario_id = %s AND MONTH(fecha) = %s AND YEAR(fecha) = %s
                """, (user_id, mes, anio_actual))
            
            row = cursor.fetchone()
            ingresos_por_mes.append(float(row['ingresos'] or 0))
            gastos_por_mes.append(float(row['gastos'] or 0))
        
        # Evolución del ahorro acumulado
        saldo_acumulado = 0
        evolucion_ahorro = []
        
        for mes in range(1, 13):
            if os.environ.get('RENDER'):
                cursor.execute("""
                    SELECT COALESCE(SUM(CASE WHEN tipo = 'ingreso' THEN monto ELSE 0 END), 0) as ingresos,
                           COALESCE(SUM(CASE WHEN tipo = 'gasto' THEN monto ELSE 0 END), 0) as gastos
                    FROM movimientos
                    WHERE usuario_id = %s AND EXTRACT(MONTH FROM fecha) <= %s AND EXTRACT(YEAR FROM fecha) = %s
                """, (user_id, mes, anio_actual))
            else:
                cursor.execute("""
                    SELECT COALESCE(SUM(CASE WHEN tipo = 'ingreso' THEN monto ELSE 0 END), 0) as ingresos,
                           COALESCE(SUM(CASE WHEN tipo = 'gasto' THEN monto ELSE 0 END), 0) as gastos
                    FROM movimientos
                    WHERE usuario_id = %s AND MONTH(fecha) <= %s AND YEAR(fecha) = %s
                """, (user_id, mes, anio_actual))
            
            row = cursor.fetchone()
            saldo_acumulado = float(row['ingresos'] or 0) - float(row['gastos'] or 0)
            evolucion_ahorro.append(saldo_acumulado)
        
        # Gastos hormiga (gastos < $100)
        if os.environ.get('RENDER'):
            cursor.execute("""
                SELECT c.nombre, COUNT(*) as frecuencia, SUM(m.monto) as total
                FROM movimientos m
                JOIN categorias c ON m.categoria_id = c.id
                WHERE m.usuario_id = %s AND m.tipo = 'gasto' AND m.monto < 100
                GROUP BY c.id, c.nombre
                ORDER BY frecuencia DESC
                LIMIT 5
            """, (user_id,))
        else:
            cursor.execute("""
                SELECT c.nombre, COUNT(*) as frecuencia, SUM(m.monto) as total
                FROM movimientos m
                JOIN categorias c ON m.categoria_id = c.id
                WHERE m.usuario_id = %s AND m.tipo = 'gasto' AND m.monto < 100
                GROUP BY c.id, c.nombre
                ORDER BY frecuencia DESC
                LIMIT 5
            """, (user_id,))
        
        gastos_hormiga = cursor.fetchall()
        cursor.close()
        conn.close()
        
        return {
            'meses': meses,
            'ingresos_mensuales': ingresos_por_mes,
            'gastos_mensuales': gastos_por_mes,
            'evolucion_ahorro': evolucion_ahorro,
            'gastos_hormiga': [
                {'categoria': g['nombre'], 'frecuencia': g['frecuencia'], 'total': float(g['total'])} 
                for g in gastos_hormiga
            ]
        }
    except Exception as e:
        print(f"Error en datos_graficos: {e}")
        if conn:
            conn.close()
        return {'error': str(e)}

# ==================== REPORTE PDF ====================

@app.route('/reporte-pdf')
@login_required
def reporte_pdf():
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from io import BytesIO
    from flask import make_response
    from calendar import monthrange
    
    conn = get_db_connection()
    if not conn:
        flash('Error de conexión', 'danger')
        return redirect(url_for('dashboard'))
    
    try:
        cursor = get_dict_cursor(conn)
        user_currency = session.get('default_currency', 'USD')
        
        cursor.execute("SELECT nombre, email, telefono, fecha_registro FROM users WHERE id = %s", (session['user_id'],))
        usuario = cursor.fetchone()
        
        mes_actual = datetime.now().month
        anio_actual = datetime.now().year
        mes_nombre = datetime(2000, mes_actual, 1).strftime('%B')
        
        if os.environ.get('RENDER'):
            cursor.execute("""
                SELECT 
                    COALESCE(SUM(CASE WHEN tipo = 'ingreso' THEN monto ELSE 0 END), 0) as total_ingresos,
                    COALESCE(SUM(CASE WHEN tipo = 'gasto' THEN monto ELSE 0 END), 0) as total_gastos,
                    COUNT(*) as total_transacciones
                FROM movimientos 
                WHERE usuario_id = %s 
                AND EXTRACT(MONTH FROM fecha) = %s AND EXTRACT(YEAR FROM fecha) = %s
            """, (session['user_id'], mes_actual, anio_actual))
        else:
            cursor.execute("""
                SELECT 
                    COALESCE(SUM(CASE WHEN tipo = 'ingreso' THEN monto ELSE 0 END), 0) as total_ingresos,
                    COALESCE(SUM(CASE WHEN tipo = 'gasto' THEN monto ELSE 0 END), 0) as total_gastos,
                    COUNT(*) as total_transacciones
                FROM movimientos 
                WHERE usuario_id = %s 
                AND MONTH(fecha) = %s AND YEAR(fecha) = %s
            """, (session['user_id'], mes_actual, anio_actual))
        resumen_mes = cursor.fetchone()
        balance = resumen_mes['total_ingresos'] - resumen_mes['total_gastos']
        
        dias_mes = monthrange(anio_actual, mes_actual)[1]
        gasto_diario = resumen_mes['total_gastos'] / dias_mes if dias_mes > 0 else 0
        
        mes_anterior = mes_actual - 1 if mes_actual > 1 else 12
        anio_anterior = anio_actual if mes_actual > 1 else anio_actual - 1
        
        if os.environ.get('RENDER'):
            cursor.execute("""
                SELECT 
                    COALESCE(SUM(CASE WHEN tipo = 'ingreso' THEN monto ELSE 0 END), 0) as total_ingresos,
                    COALESCE(SUM(CASE WHEN tipo = 'gasto' THEN monto ELSE 0 END), 0) as total_gastos
                FROM movimientos 
                WHERE usuario_id = %s 
                AND EXTRACT(MONTH FROM fecha) = %s AND EXTRACT(YEAR FROM fecha) = %s
            """, (session['user_id'], mes_anterior, anio_anterior))
        else:
            cursor.execute("""
                SELECT 
                    COALESCE(SUM(CASE WHEN tipo = 'ingreso' THEN monto ELSE 0 END), 0) as total_ingresos,
                    COALESCE(SUM(CASE WHEN tipo = 'gasto' THEN monto ELSE 0 END), 0) as total_gastos
                FROM movimientos 
                WHERE usuario_id = %s 
                AND MONTH(fecha) = %s AND YEAR(fecha) = %s
            """, (session['user_id'], mes_anterior, anio_anterior))
        mes_anterior_data = cursor.fetchone()
        
        variacion_ingresos = ((resumen_mes['total_ingresos'] - mes_anterior_data['total_ingresos']) / mes_anterior_data['total_ingresos'] * 100) if mes_anterior_data['total_ingresos'] > 0 else 0
        variacion_gastos = ((resumen_mes['total_gastos'] - mes_anterior_data['total_gastos']) / mes_anterior_data['total_gastos'] * 100) if mes_anterior_data['total_gastos'] > 0 else 0
        
        # Gastos por categoría
        if os.environ.get('RENDER'):
            cursor.execute("""
                SELECT c.nombre as categoria, COALESCE(SUM(m.monto), 0) as total
                FROM categorias c
                LEFT JOIN movimientos m ON c.id = m.categoria_id 
                    AND m.tipo = 'gasto' AND m.usuario_id = %s
                    AND EXTRACT(MONTH FROM m.fecha) = %s AND EXTRACT(YEAR FROM m.fecha) = %s
                GROUP BY c.id, c.nombre
                HAVING COALESCE(SUM(m.monto), 0) > 0
                ORDER BY total DESC
            """, (session['user_id'], mes_actual, anio_actual))
        else:
            cursor.execute("""
                SELECT c.nombre as categoria, COALESCE(SUM(m.monto), 0) as total
                FROM categorias c
                LEFT JOIN movimientos m ON c.id = m.categoria_id 
                    AND m.tipo = 'gasto' AND m.usuario_id = %s
                    AND MONTH(m.fecha) = %s AND YEAR(m.fecha) = %s
                GROUP BY c.id, c.nombre
                HAVING total > 0
                ORDER BY total DESC
            """, (session['user_id'], mes_actual, anio_actual))
        gastos_categoria = cursor.fetchall()
        
        # Ingresos por categoría
        if os.environ.get('RENDER'):
            cursor.execute("""
                SELECT c.nombre as categoria, COALESCE(SUM(m.monto), 0) as total
                FROM categorias c
                LEFT JOIN movimientos m ON c.id = m.categoria_id 
                    AND m.tipo = 'ingreso' AND m.usuario_id = %s
                    AND EXTRACT(MONTH FROM m.fecha) = %s AND EXTRACT(YEAR FROM m.fecha) = %s
                GROUP BY c.id, c.nombre
                HAVING COALESCE(SUM(m.monto), 0) > 0
                ORDER BY total DESC
            """, (session['user_id'], mes_actual, anio_actual))
        else:
            cursor.execute("""
                SELECT c.nombre as categoria, COALESCE(SUM(m.monto), 0) as total
                FROM categorias c
                LEFT JOIN movimientos m ON c.id = m.categoria_id 
                    AND m.tipo = 'ingreso' AND m.usuario_id = %s
                    AND MONTH(m.fecha) = %s AND YEAR(m.fecha) = %s
                GROUP BY c.id, c.nombre
                HAVING total > 0
                ORDER BY total DESC
            """, (session['user_id'], mes_actual, anio_actual))
        ingresos_categoria = cursor.fetchall()
        
        # Presupuestos
        if os.environ.get('RENDER'):
            cursor.execute("""
                SELECT c.nombre as categoria, p.limite, 
                       COALESCE(SUM(m.monto), 0) as gastado,
                       (COALESCE(SUM(m.monto), 0) / p.limite * 100) as porcentaje
                FROM presupuestos p
                JOIN categorias c ON p.categoria_id = c.id
                LEFT JOIN movimientos m ON c.id = m.categoria_id 
                    AND m.tipo = 'gasto' AND m.usuario_id = %s
                    AND EXTRACT(MONTH FROM m.fecha) = %s AND EXTRACT(YEAR FROM m.fecha) = %s
                WHERE p.usuario_id = %s AND p.mes = %s AND p.anio = %s
                GROUP BY c.id, c.nombre, p.limite
            """, (session['user_id'], mes_actual, anio_actual, session['user_id'], mes_actual, anio_actual))
        else:
            cursor.execute("""
                SELECT c.nombre as categoria, p.limite, 
                       COALESCE(SUM(m.monto), 0) as gastado,
                       (COALESCE(SUM(m.monto), 0) / p.limite * 100) as porcentaje
                FROM presupuestos p
                JOIN categorias c ON p.categoria_id = c.id
                LEFT JOIN movimientos m ON c.id = m.categoria_id 
                    AND m.tipo = 'gasto' AND m.usuario_id = %s
                    AND MONTH(m.fecha) = %s AND YEAR(m.fecha) = %s
                WHERE p.usuario_id = %s AND p.mes = %s AND p.anio = %s
                GROUP BY c.id, c.nombre, p.limite
            """, (session['user_id'], mes_actual, anio_actual, session['user_id'], mes_actual, anio_actual))
        presupuestos = cursor.fetchall()
        
        # Metas - obtener y procesar en Python
        cursor.execute("""
            SELECT nombre, monto_objetivo, monto_actual, fecha_limite
            FROM metas WHERE usuario_id = %s
        """, (session['user_id'],))
        metas = cursor.fetchall()
        
        # Procesar metas en Python para evitar problemas de compatibilidad
        for meta in metas:
            if meta['monto_objetivo'] and meta['monto_objetivo'] > 0:
                meta['porcentaje'] = (meta['monto_actual'] / meta['monto_objetivo'] * 100)
            else:
                meta['porcentaje'] = 0
        
        # Evolución últimos 6 meses
        evolucion = []
        for i in range(5, -1, -1):
            mes_num = mes_actual - i
            año_num = anio_actual
            if mes_num <= 0:
                mes_num += 12
                año_num -= 1
            
            if os.environ.get('RENDER'):
                cursor.execute("""
                    SELECT 
                        COALESCE(SUM(CASE WHEN tipo = 'ingreso' THEN monto ELSE 0 END), 0) as ingresos,
                        COALESCE(SUM(CASE WHEN tipo = 'gasto' THEN monto ELSE 0 END), 0) as gastos
                    FROM movimientos
                    WHERE usuario_id = %s AND EXTRACT(MONTH FROM fecha) = %s AND EXTRACT(YEAR FROM fecha) = %s
                """, (session['user_id'], mes_num, año_num))
            else:
                cursor.execute("""
                    SELECT 
                        COALESCE(SUM(CASE WHEN tipo = 'ingreso' THEN monto ELSE 0 END), 0) as ingresos,
                        COALESCE(SUM(CASE WHEN tipo = 'gasto' THEN monto ELSE 0 END), 0) as gastos
                    FROM movimientos
                    WHERE usuario_id = %s AND MONTH(fecha) = %s AND YEAR(fecha) = %s
                """, (session['user_id'], mes_num, año_num))
            data = cursor.fetchone()
            evolucion.append({
                'mes': datetime(2000, mes_num, 1).strftime('%b %Y'),
                'ingresos': data['ingresos'],
                'gastos': data['gastos'],
                'balance': data['ingresos'] - data['gastos']
            })
        
        # Últimos movimientos
        if os.environ.get('RENDER'):
            cursor.execute("""
                SELECT m.tipo, m.monto, m.descripcion, TO_CHAR(m.fecha, 'DD/MM/YYYY') as fecha, c.nombre as categoria
                FROM movimientos m
                JOIN categorias c ON m.categoria_id = c.id
                WHERE m.usuario_id = %s
                ORDER BY m.fecha DESC
                LIMIT 20
            """, (session['user_id'],))
        else:
            cursor.execute("""
                SELECT m.tipo, m.monto, m.descripcion, DATE_FORMAT(m.fecha, '%d/%m/%Y') as fecha, c.nombre as categoria
                FROM movimientos m
                JOIN categorias c ON m.categoria_id = c.id
                WHERE m.usuario_id = %s
                ORDER BY m.fecha DESC
                LIMIT 20
            """, (session['user_id'],))
        ultimos_movimientos = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        # Generar PDF - (el código de generación de PDF es extenso, pero se mantiene igual)
        # ... [código de generación de PDF aquí] ...
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        styles = getSampleStyleSheet()
        story = []
        
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=24, textColor=colors.HexColor('#1E293B'), alignment=1, spaceAfter=20)
        section_style = ParagraphStyle('Section', parent=styles['Heading2'], fontSize=16, textColor=colors.HexColor('#3B82F6'), spaceAfter=15, spaceBefore=20)
        
        story.append(Paragraph("FinanTrack - Reporte Financiero", title_style))
        story.append(Paragraph(f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
        story.append(Paragraph(f"Moneda: {user_currency}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Información del usuario
        info_data = [['Información del Usuario', '']]
        info_data.append(['Nombre', usuario['nombre'] or 'No registrado'])
        info_data.append(['Email', usuario['email']])
        info_data.append(['Teléfono', usuario['telefono'] if usuario['telefono'] else 'No registrado'])
        info_data.append(['Miembro desde', usuario['fecha_registro'].strftime('%d/%m/%Y') if usuario['fecha_registro'] else 'N/A'])
        
        info_table = Table(info_data, colWidths=[150, 250])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3B82F6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8FAFC')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 20))
        
        # Resumen del mes
        story.append(Paragraph(f"Resumen del mes - {mes_nombre} {anio_actual}", section_style))
        
        resumen_data = [
            ['Concepto', 'Monto', 'vs mes anterior'],
            ['Total Ingresos', f'{user_currency} {resumen_mes["total_ingresos"]:,.2f}', f'{variacion_ingresos:+.1f}%'],
            ['Total Gastos', f'{user_currency} {resumen_mes["total_gastos"]:,.2f}', f'{variacion_gastos:+.1f}%'],
            ['Balance', f'{user_currency} {balance:,.2f}', ''],
            ['Gasto diario promedio', f'{user_currency} {gasto_diario:,.2f}', ''],
            ['Total transacciones', str(resumen_mes['total_transacciones']), '']
        ]
        
        resumen_table = Table(resumen_data, colWidths=[150, 120, 120])
        resumen_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#34D399')),
            ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#F87171')),
            ('BACKGROUND', (0, 3), (-1, 3), colors.HexColor('#60A5FA')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
            ('ALIGN', (1, 1), (2, -1), 'RIGHT'),
        ]))
        story.append(resumen_table)
        story.append(Spacer(1, 20))
        
        # Gastos por categoría
        if gastos_categoria:
            story.append(Paragraph("Gastos por Categoría", section_style))
            total_gastos_cat = sum(g['total'] for g in gastos_categoria)
            gastos_data = [['Categoría', 'Monto', '% del total']]
            for g in gastos_categoria:
                porcentaje = (g['total'] / total_gastos_cat * 100) if total_gastos_cat > 0 else 0
                gastos_data.append([g['categoria'], f'{user_currency} {g["total"]:,.2f}', f'{porcentaje:.1f}%'])
            
            gastos_table = Table(gastos_data, colWidths=[180, 120, 100])
            gastos_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F87171')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#FECACA')),
                ('ALIGN', (1, 1), (2, -1), 'RIGHT'),
            ]))
            story.append(gastos_table)
            story.append(Spacer(1, 15))
        
        # Presupuestos
        if presupuestos:
            story.append(Paragraph("Estado de Presupuestos", section_style))
            presupuestos_data = [['Categoría', 'Límite', 'Gastado', 'Restante', '%']]
            for p in presupuestos:
                restante = p['limite'] - p['gastado']
                presupuestos_data.append([p['categoria'], f'{user_currency} {p["limite"]:,.2f}', f'{user_currency} {p["gastado"]:,.2f}', f'{user_currency} {restante:,.2f}', f'{p["porcentaje"]:.1f}%'])
            
            presupuestos_table = Table(presupuestos_data, colWidths=[120, 100, 100, 100, 70])
            presupuestos_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8B5CF6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
                ('ALIGN', (1, 1), (4, -1), 'RIGHT'),
            ]))
            story.append(presupuestos_table)
            story.append(Spacer(1, 15))
        
        # Metas
        if metas:
            story.append(Paragraph("Metas de Ahorro", section_style))
            metas_data = [['Meta', 'Objetivo', 'Ahorrado', 'Progreso', 'Fecha límite']]
            for m in metas:
                fecha_limite = m['fecha_limite'].strftime('%d/%m/%Y') if m['fecha_limite'] else 'Sin fecha'
                metas_data.append([m['nombre'], f'{user_currency} {m["monto_objetivo"]:,.2f}', f'{user_currency} {m["monto_actual"]:,.2f}', f'{m["porcentaje"]:.1f}%', fecha_limite])
            
            metas_table = Table(metas_data, colWidths=[150, 100, 100, 80, 100])
            metas_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F59E0B')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
                ('ALIGN', (1, 1), (3, -1), 'RIGHT'),
            ]))
            story.append(metas_table)
            story.append(Spacer(1, 15))
        
        # Evolución
        story.append(Paragraph("Evolución últimos 6 meses", section_style))
        evolucion_data = [['Mes', 'Ingresos', 'Gastos', 'Balance']]
        for ev in evolucion:
            evolucion_data.append([ev['mes'], f'{user_currency} {ev["ingresos"]:,.2f}', f'{user_currency} {ev["gastos"]:,.2f}', f'{user_currency} {ev["balance"]:,.2f}'])
        
        evolucion_table = Table(evolucion_data, colWidths=[100, 100, 100, 100])
        evolucion_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3B82F6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
            ('ALIGN', (1, 1), (3, -1), 'RIGHT'),
        ]))
        story.append(evolucion_table)
        story.append(Spacer(1, 20))
        
        # Últimos movimientos
        if ultimos_movimientos:
            story.append(PageBreak())
            story.append(Paragraph("Últimos Movimientos", section_style))
            movimientos_data = [['Fecha', 'Tipo', 'Categoría', 'Descripción', 'Monto']]
            for m in ultimos_movimientos:
                desc = (m['descripcion'][:40] + '...') if m['descripcion'] and len(m['descripcion']) > 40 else (m['descripcion'] or '-')
                movimientos_data.append([m['fecha'], 'Ingreso' if m['tipo'] == 'ingreso' else 'Gasto', m['categoria'], desc, f'{user_currency} {m["monto"]:,.2f}'])
            
            movimientos_table = Table(movimientos_data, colWidths=[80, 60, 90, 140, 80])
            movimientos_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
                ('ALIGN', (4, 1), (4, -1), 'RIGHT'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#F8FAFC'), colors.HexColor('#FFFFFF')]),
            ]))
            story.append(movimientos_table)
        
        # Consejo
        story.append(Spacer(1, 30))
        consejo_texto = ""
        if balance > 500:
            consejo_texto = "¡Excelente trabajo financiero! Has logrado un ahorro significativo. Considera invertir una parte para hacer crecer tu dinero."
        elif balance > 0:
            consejo_texto = f"Vas por buen camino con un ahorro de {user_currency} {balance:,.2f}. Intenta mantener este hábito."
        elif balance == 0:
            consejo_texto = "Tus ingresos igualan a tus gastos. Busca pequeñas formas de ahorrar para construir un colchón financiero."
        else:
            consejo_texto = f"Tus gastos superan tus ingresos por {user_currency} {abs(balance):,.2f}. Revisa tus gastos hormiga y busca reducir gastos innecesarios."
        
        consejo_style = ParagraphStyle('Consejo', parent=styles['Normal'], fontSize=11, textColor=colors.HexColor('#3B82F6'), alignment=1, backColor=colors.HexColor('#EFF6FF'), spaceAfter=10, spaceBefore=10)
        story.append(Paragraph(consejo_texto, consejo_style))
        
        doc.build(story)
        buffer.seek(0)
        
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=finantrack_reporte_{datetime.now().strftime("%Y%m%d")}.pdf'
        
        return response
        
    except Exception as e:
        print(f"Error generando PDF: {e}")
        import traceback
        traceback.print_exc()
        if conn:
            conn.close()
        flash('Error al generar el reporte PDF', 'danger')
        return redirect(url_for('dashboard'))

# ==================== REPORTE EXCEL ====================

@app.route('/reporte-excel')
@login_required
def reporte_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from flask import make_response
    from calendar import monthrange
    
    conn = get_db_connection()
    if not conn:
        flash('Error de conexión', 'danger')
        return redirect(url_for('dashboard'))
    
    try:
        cursor = get_dict_cursor(conn)
        user_currency = session.get('default_currency', 'USD')
        
        cursor.execute("SELECT nombre, email, telefono, fecha_registro FROM users WHERE id = %s", (session['user_id'],))
        usuario = cursor.fetchone()
        
        mes_actual = datetime.now().month
        anio_actual = datetime.now().year
        mes_nombre = datetime(2000, mes_actual, 1).strftime('%B')
        
        if os.environ.get('RENDER'):
            cursor.execute("""
                SELECT 
                    COALESCE(SUM(CASE WHEN tipo = 'ingreso' THEN monto ELSE 0 END), 0) as total_ingresos,
                    COALESCE(SUM(CASE WHEN tipo = 'gasto' THEN monto ELSE 0 END), 0) as total_gastos,
                    COUNT(*) as total_transacciones
                FROM movimientos 
                WHERE usuario_id = %s 
                AND EXTRACT(MONTH FROM fecha) = %s AND EXTRACT(YEAR FROM fecha) = %s
            """, (session['user_id'], mes_actual, anio_actual))
        else:
            cursor.execute("""
                SELECT 
                    COALESCE(SUM(CASE WHEN tipo = 'ingreso' THEN monto ELSE 0 END), 0) as total_ingresos,
                    COALESCE(SUM(CASE WHEN tipo = 'gasto' THEN monto ELSE 0 END), 0) as total_gastos,
                    COUNT(*) as total_transacciones
                FROM movimientos 
                WHERE usuario_id = %s 
                AND MONTH(fecha) = %s AND YEAR(fecha) = %s
            """, (session['user_id'], mes_actual, anio_actual))
        resumen_mes = cursor.fetchone()
        balance = resumen_mes['total_ingresos'] - resumen_mes['total_gastos']
        
        dias_mes = monthrange(anio_actual, mes_actual)[1]
        gasto_diario = resumen_mes['total_gastos'] / dias_mes if dias_mes > 0 else 0
        
        mes_anterior = mes_actual - 1 if mes_actual > 1 else 12
        anio_anterior = anio_actual if mes_actual > 1 else anio_actual - 1
        
        if os.environ.get('RENDER'):
            cursor.execute("""
                SELECT 
                    COALESCE(SUM(CASE WHEN tipo = 'ingreso' THEN monto ELSE 0 END), 0) as total_ingresos,
                    COALESCE(SUM(CASE WHEN tipo = 'gasto' THEN monto ELSE 0 END), 0) as total_gastos
                FROM movimientos 
                WHERE usuario_id = %s 
                AND EXTRACT(MONTH FROM fecha) = %s AND EXTRACT(YEAR FROM fecha) = %s
            """, (session['user_id'], mes_anterior, anio_anterior))
        else:
            cursor.execute("""
                SELECT 
                    COALESCE(SUM(CASE WHEN tipo = 'ingreso' THEN monto ELSE 0 END), 0) as total_ingresos,
                    COALESCE(SUM(CASE WHEN tipo = 'gasto' THEN monto ELSE 0 END), 0) as total_gastos
                FROM movimientos 
                WHERE usuario_id = %s 
                AND MONTH(fecha) = %s AND YEAR(fecha) = %s
            """, (session['user_id'], mes_anterior, anio_anterior))
        mes_anterior_data = cursor.fetchone()
        
        # Gastos por categoría
        if os.environ.get('RENDER'):
            cursor.execute("""
                SELECT c.nombre as categoria, COALESCE(SUM(m.monto), 0) as total
                FROM categorias c
                LEFT JOIN movimientos m ON c.id = m.categoria_id 
                    AND m.tipo = 'gasto' AND m.usuario_id = %s
                    AND EXTRACT(MONTH FROM m.fecha) = %s AND EXTRACT(YEAR FROM m.fecha) = %s
                GROUP BY c.id, c.nombre
                HAVING COALESCE(SUM(m.monto), 0) > 0
                ORDER BY total DESC
            """, (session['user_id'], mes_actual, anio_actual))
        else:
            cursor.execute("""
                SELECT c.nombre as categoria, COALESCE(SUM(m.monto), 0) as total
                FROM categorias c
                LEFT JOIN movimientos m ON c.id = m.categoria_id 
                    AND m.tipo = 'gasto' AND m.usuario_id = %s
                    AND MONTH(m.fecha) = %s AND YEAR(m.fecha) = %s
                GROUP BY c.id, c.nombre
                HAVING total > 0
                ORDER BY total DESC
            """, (session['user_id'], mes_actual, anio_actual))
        gastos_categoria = cursor.fetchall()
        
        # Ingresos por categoría
        if os.environ.get('RENDER'):
            cursor.execute("""
                SELECT c.nombre as categoria, COALESCE(SUM(m.monto), 0) as total
                FROM categorias c
                LEFT JOIN movimientos m ON c.id = m.categoria_id 
                    AND m.tipo = 'ingreso' AND m.usuario_id = %s
                    AND EXTRACT(MONTH FROM m.fecha) = %s AND EXTRACT(YEAR FROM m.fecha) = %s
                GROUP BY c.id, c.nombre
                HAVING COALESCE(SUM(m.monto), 0) > 0
                ORDER BY total DESC
            """, (session['user_id'], mes_actual, anio_actual))
        else:
            cursor.execute("""
                SELECT c.nombre as categoria, COALESCE(SUM(m.monto), 0) as total
                FROM categorias c
                LEFT JOIN movimientos m ON c.id = m.categoria_id 
                    AND m.tipo = 'ingreso' AND m.usuario_id = %s
                    AND MONTH(m.fecha) = %s AND YEAR(m.fecha) = %s
                GROUP BY c.id, c.nombre
                HAVING total > 0
                ORDER BY total DESC
            """, (session['user_id'], mes_actual, anio_actual))
        ingresos_categoria = cursor.fetchall()
        
        # Presupuestos
        if os.environ.get('RENDER'):
            cursor.execute("""
                SELECT c.nombre as categoria, p.limite, 
                       COALESCE(SUM(m.monto), 0) as gastado,
                       (COALESCE(SUM(m.monto), 0) / p.limite * 100) as porcentaje
                FROM presupuestos p
                JOIN categorias c ON p.categoria_id = c.id
                LEFT JOIN movimientos m ON c.id = m.categoria_id 
                    AND m.tipo = 'gasto' AND m.usuario_id = %s
                    AND EXTRACT(MONTH FROM m.fecha) = %s AND EXTRACT(YEAR FROM m.fecha) = %s
                WHERE p.usuario_id = %s AND p.mes = %s AND p.anio = %s
                GROUP BY c.id, c.nombre, p.limite
            """, (session['user_id'], mes_actual, anio_actual, session['user_id'], mes_actual, anio_actual))
        else:
            cursor.execute("""
                SELECT c.nombre as categoria, p.limite, 
                       COALESCE(SUM(m.monto), 0) as gastado,
                       (COALESCE(SUM(m.monto), 0) / p.limite * 100) as porcentaje
                FROM presupuestos p
                JOIN categorias c ON p.categoria_id = c.id
                LEFT JOIN movimientos m ON c.id = m.categoria_id 
                    AND m.tipo = 'gasto' AND m.usuario_id = %s
                    AND MONTH(m.fecha) = %s AND YEAR(m.fecha) = %s
                WHERE p.usuario_id = %s AND p.mes = %s AND p.anio = %s
                GROUP BY c.id, c.nombre, p.limite
            """, (session['user_id'], mes_actual, anio_actual, session['user_id'], mes_actual, anio_actual))
        presupuestos = cursor.fetchall()
        
        # Metas
        cursor.execute("""
            SELECT nombre, monto_objetivo, monto_actual, fecha_limite
            FROM metas WHERE usuario_id = %s
        """, (session['user_id'],))
        metas = cursor.fetchall()
        
        # Procesar metas en Python
        for meta in metas:
            if meta['monto_objetivo'] and meta['monto_objetivo'] > 0:
                meta['porcentaje'] = (meta['monto_actual'] / meta['monto_objetivo'] * 100)
            else:
                meta['porcentaje'] = 0
        
        # Evolución últimos 6 meses
        evolucion = []
        for i in range(5, -1, -1):
            mes_num = mes_actual - i
            año_num = anio_actual
            if mes_num <= 0:
                mes_num += 12
                año_num -= 1
            
            if os.environ.get('RENDER'):
                cursor.execute("""
                    SELECT 
                        COALESCE(SUM(CASE WHEN tipo = 'ingreso' THEN monto ELSE 0 END), 0) as ingresos,
                        COALESCE(SUM(CASE WHEN tipo = 'gasto' THEN monto ELSE 0 END), 0) as gastos
                    FROM movimientos
                    WHERE usuario_id = %s AND EXTRACT(MONTH FROM fecha) = %s AND EXTRACT(YEAR FROM fecha) = %s
                """, (session['user_id'], mes_num, año_num))
            else:
                cursor.execute("""
                    SELECT 
                        COALESCE(SUM(CASE WHEN tipo = 'ingreso' THEN monto ELSE 0 END), 0) as ingresos,
                        COALESCE(SUM(CASE WHEN tipo = 'gasto' THEN monto ELSE 0 END), 0) as gastos
                    FROM movimientos
                    WHERE usuario_id = %s AND MONTH(fecha) = %s AND YEAR(fecha) = %s
                """, (session['user_id'], mes_num, año_num))
            data = cursor.fetchone()
            evolucion.append({
                'mes': datetime(2000, mes_num, 1).strftime('%b %Y'),
                'ingresos': data['ingresos'],
                'gastos': data['gastos'],
                'balance': data['ingresos'] - data['gastos']
            })
        
        # Todos los movimientos
        if os.environ.get('RENDER'):
            cursor.execute("""
                SELECT m.id, m.tipo, m.monto, m.descripcion, TO_CHAR(m.fecha, 'DD/MM/YYYY') as fecha, c.nombre as categoria
                FROM movimientos m
                JOIN categorias c ON m.categoria_id = c.id
                WHERE m.usuario_id = %s
                ORDER BY m.fecha DESC
            """, (session['user_id'],))
        else:
            cursor.execute("""
                SELECT m.id, m.tipo, m.monto, m.descripcion, DATE_FORMAT(m.fecha, '%d/%m/%Y') as fecha, c.nombre as categoria
                FROM movimientos m
                JOIN categorias c ON m.categoria_id = c.id
                WHERE m.usuario_id = %s
                ORDER BY m.fecha DESC
            """, (session['user_id'],))
        todos_movimientos = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        # Crear libro de Excel - (el código de generación de Excel es extenso, pero se mantiene igual)
        wb = openpyxl.Workbook()
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        title_font = Font(bold=True, size=14)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        green_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        
        # Hoja 1: Resumen General
        ws_resumen = wb.active
        ws_resumen.title = "Resumen General"
        
        ws_resumen['A1'] = f"FINANTRACK - REPORTE FINANCIERO COMPLETO ({user_currency})"
        ws_resumen['A1'].font = title_font
        ws_resumen.merge_cells('A1:F1')
        
        ws_resumen['A3'] = "Información del Usuario"
        ws_resumen['A3'].font = Font(bold=True)
        ws_resumen['A4'] = "Nombre:"
        ws_resumen['B4'] = usuario['nombre'] or 'No registrado'
        ws_resumen['A5'] = "Email:"
        ws_resumen['B5'] = usuario['email']
        ws_resumen['A6'] = "Teléfono:"
        ws_resumen['B6'] = usuario['telefono'] if usuario['telefono'] else 'No registrado'
        ws_resumen['A7'] = "Miembro desde:"
        ws_resumen['B7'] = usuario['fecha_registro'].strftime('%d/%m/%Y') if usuario['fecha_registro'] else 'N/A'
        ws_resumen['A8'] = "Moneda:"
        ws_resumen['B8'] = user_currency
        ws_resumen['A9'] = "Reporte generado:"
        ws_resumen['B9'] = datetime.now().strftime('%d/%m/%Y %H:%M')
        
        ws_resumen['A11'] = f"Resumen del mes - {mes_nombre} {anio_actual}"
        ws_resumen['A11'].font = Font(bold=True)
        
        resumen_data = [
            ['Total Ingresos', f'{user_currency} {resumen_mes["total_ingresos"]:,.2f}'],
            ['Total Gastos', f'{user_currency} {resumen_mes["total_gastos"]:,.2f}'],
            ['Balance', f'{user_currency} {balance:,.2f}'],
            ['Gasto diario promedio', f'{user_currency} {gasto_diario:,.2f}'],
            ['Total transacciones', resumen_mes['total_transacciones']],
            ['vs mes anterior ingresos', f'{((resumen_mes["total_ingresos"] - mes_anterior_data["total_ingresos"]) / mes_anterior_data["total_ingresos"] * 100):+.1f}%' if mes_anterior_data["total_ingresos"] > 0 else 'N/A'],
            ['vs mes anterior gastos', f'{((resumen_mes["total_gastos"] - mes_anterior_data["total_gastos"]) / mes_anterior_data["total_gastos"] * 100):+.1f}%' if mes_anterior_data["total_gastos"] > 0 else 'N/A']
        ]
        
        for i, row in enumerate(resumen_data, start=12):
            ws_resumen[f'A{i}'] = row[0]
            ws_resumen[f'B{i}'] = row[1]
            ws_resumen[f'A{i}'].font = Font(bold=True)
        
        ws_resumen['A20'] = "Evolución últimos 6 meses"
        ws_resumen['A20'].font = Font(bold=True)
        
        headers_evolucion = ['Mes', 'Ingresos', 'Gastos', 'Balance']
        for col, header in enumerate(headers_evolucion, 1):
            cell = ws_resumen.cell(row=21, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        for row, ev in enumerate(evolucion, 22):
            ws_resumen.cell(row=row, column=1, value=ev['mes'])
            ws_resumen.cell(row=row, column=2, value=ev['ingresos'])
            ws_resumen.cell(row=row, column=3, value=ev['gastos'])
            ws_resumen.cell(row=row, column=4, value=ev['balance'])
            for col in range(1, 5):
                ws_resumen.cell(row=row, column=col).border = border
                if col in [2, 3, 4]:
                    ws_resumen.cell(row=row, column=col).number_format = f'"{user_currency} "#,##0.00'
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws_resumen.column_dimensions[col].width = 25
        
        # Hoja 2: Gastos por Categoría
        ws_gastos = wb.create_sheet("Gastos por Categoría")
        ws_gastos['A1'] = f"GASTOS POR CATEGORÍA ({user_currency})"
        ws_gastos['A1'].font = title_font
        ws_gastos.merge_cells('A1:D1')
        
        headers_gastos = ['Categoría', 'Monto', '% del total']
        for col, header in enumerate(headers_gastos, 1):
            cell = ws_gastos.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        total_gastos_cat = sum(g['total'] for g in gastos_categoria)
        for row, g in enumerate(gastos_categoria, 4):
            porcentaje = (g['total'] / total_gastos_cat * 100) if total_gastos_cat > 0 else 0
            ws_gastos.cell(row=row, column=1, value=g['categoria'])
            ws_gastos.cell(row=row, column=2, value=g['total'])
            ws_gastos.cell(row=row, column=3, value=porcentaje)
            for col in range(1, 4):
                ws_gastos.cell(row=row, column=col).border = border
                if col == 2:
                    ws_gastos.cell(row=row, column=col).number_format = f'"{user_currency} "#,##0.00'
                elif col == 3:
                    ws_gastos.cell(row=row, column=col).number_format = '0.00"%'
        
        for col in ['A', 'B', 'C']:
            ws_gastos.column_dimensions[col].width = 20
        
        # Hoja 3: Ingresos por Categoría
        ws_ingresos = wb.create_sheet("Ingresos por Categoría")
        ws_ingresos['A1'] = f"INGRESOS POR CATEGORÍA ({user_currency})"
        ws_ingresos['A1'].font = title_font
        ws_ingresos.merge_cells('A1:C1')
        
        headers_ingresos = ['Categoría', 'Monto', '% del total']
        for col, header in enumerate(headers_ingresos, 1):
            cell = ws_ingresos.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        total_ingresos_cat = sum(i['total'] for i in ingresos_categoria)
        for row, i in enumerate(ingresos_categoria, 4):
            porcentaje = (i['total'] / total_ingresos_cat * 100) if total_ingresos_cat > 0 else 0
            ws_ingresos.cell(row=row, column=1, value=i['categoria'])
            ws_ingresos.cell(row=row, column=2, value=i['total'])
            ws_ingresos.cell(row=row, column=3, value=porcentaje)
            for col in range(1, 4):
                ws_ingresos.cell(row=row, column=col).border = border
                if col == 2:
                    ws_ingresos.cell(row=row, column=col).number_format = f'"{user_currency} "#,##0.00'
                elif col == 3:
                    ws_ingresos.cell(row=row, column=col).number_format = '0.00"%'
        
        for col in ['A', 'B', 'C']:
            ws_ingresos.column_dimensions[col].width = 20
        
        # Hoja 4: Presupuestos
        ws_presupuestos = wb.create_sheet("Presupuestos")
        ws_presupuestos['A1'] = f"ESTADO DE PRESUPUESTOS ({user_currency})"
        ws_presupuestos['A1'].font = title_font
        ws_presupuestos.merge_cells('A1:E1')
        
        headers_presupuestos = ['Categoría', 'Límite', 'Gastado', 'Restante', 'Progreso']
        for col, header in enumerate(headers_presupuestos, 1):
            cell = ws_presupuestos.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        for row, p in enumerate(presupuestos, 4):
            restante = p['limite'] - p['gastado']
            ws_presupuestos.cell(row=row, column=1, value=p['categoria'])
            ws_presupuestos.cell(row=row, column=2, value=p['limite'])
            ws_presupuestos.cell(row=row, column=3, value=p['gastado'])
            ws_presupuestos.cell(row=row, column=4, value=restante)
            ws_presupuestos.cell(row=row, column=5, value=p['porcentaje'])
            for col in range(1, 6):
                ws_presupuestos.cell(row=row, column=col).border = border
                if col in [2, 3, 4]:
                    ws_presupuestos.cell(row=row, column=col).number_format = f'"{user_currency} "#,##0.00'
                elif col == 5:
                    ws_presupuestos.cell(row=row, column=col).number_format = '0.00"%'
            if p['gastado'] > p['limite']:
                for col in range(1, 6):
                    ws_presupuestos.cell(row=row, column=col).fill = red_fill
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws_presupuestos.column_dimensions[col].width = 18
        
        # Hoja 5: Metas
        ws_metas = wb.create_sheet("Metas de Ahorro")
        ws_metas['A1'] = f"METAS DE AHORRO ({user_currency})"
        ws_metas['A1'].font = title_font
        ws_metas.merge_cells('A1:E1')
        
        headers_metas = ['Meta', 'Objetivo', 'Ahorrado', 'Progreso', 'Fecha límite']
        for col, header in enumerate(headers_metas, 1):
            cell = ws_metas.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        for row, m in enumerate(metas, 4):
            ws_metas.cell(row=row, column=1, value=m['nombre'])
            ws_metas.cell(row=row, column=2, value=m['monto_objetivo'])
            ws_metas.cell(row=row, column=3, value=m['monto_actual'])
            ws_metas.cell(row=row, column=4, value=m['porcentaje'])
            ws_metas.cell(row=row, column=5, value=m['fecha_limite'].strftime('%d/%m/%Y') if m['fecha_limite'] else 'Sin fecha')
            for col in range(1, 6):
                ws_metas.cell(row=row, column=col).border = border
                if col in [2, 3]:
                    ws_metas.cell(row=row, column=col).number_format = f'"{user_currency} "#,##0.00'
                elif col == 4:
                    ws_metas.cell(row=row, column=col).number_format = '0.00"%'
            if m['porcentaje'] >= 100:
                for col in range(1, 6):
                    ws_metas.cell(row=row, column=col).fill = green_fill
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws_metas.column_dimensions[col].width = 18
        
        # Hoja 6: Todos los movimientos
        ws_movimientos = wb.create_sheet("Todos los Movimientos")
        ws_movimientos['A1'] = f"REGISTRO COMPLETO DE MOVIMIENTOS ({user_currency})"
        ws_movimientos['A1'].font = title_font
        ws_movimientos.merge_cells('A1:F1')
        
        headers_movimientos = ['ID', 'Fecha', 'Tipo', 'Categoría', 'Descripción', 'Monto']
        for col, header in enumerate(headers_movimientos, 1):
            cell = ws_movimientos.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        for row, m in enumerate(todos_movimientos, 4):
            ws_movimientos.cell(row=row, column=1, value=m['id'])
            ws_movimientos.cell(row=row, column=2, value=m['fecha'])
            ws_movimientos.cell(row=row, column=3, value="Ingreso" if m['tipo'] == 'ingreso' else "Gasto")
            ws_movimientos.cell(row=row, column=4, value=m['categoria'])
            ws_movimientos.cell(row=row, column=5, value=m['descripcion'] if m['descripcion'] else '-')
            ws_movimientos.cell(row=row, column=6, value=m['monto'])
            for col in range(1, 7):
                ws_movimientos.cell(row=row, column=col).border = border
                if col == 6:
                    ws_movimientos.cell(row=row, column=col).number_format = f'"{user_currency} "#,##0.00'
            if m['tipo'] == 'ingreso':
                for col in range(1, 7):
                    ws_movimientos.cell(row=row, column=col).fill = green_fill
            else:
                for col in range(1, 7):
                    ws_movimientos.cell(row=row, column=col).fill = red_fill
        
        for col in range(1, 7):
            ws_movimientos.column_dimensions[get_column_letter(col)].width = 20
        
        # Hoja 7: Consejos
        ws_consejos = wb.create_sheet("Consejos")
        ws_consejos['A1'] = "CONSEJOS FINANCIEROS PERSONALIZADOS"
        ws_consejos['A1'].font = title_font
        ws_consejos.merge_cells('A1:B1')
        
        consejos = []
        if balance > 500:
            consejos.append(["🏆 Excelente ahorro", f"Has ahorrado {user_currency} {balance:,.2f} este mes."])
            consejos.append(["💡 Inversión recomendada", "Con este nivel de ahorro, podrías comenzar a invertir."])
        elif balance > 0:
            consejos.append(["👍 Buen trabajo", f"Has ahorrado {user_currency} {balance:,.2f} este mes."])
            consejos.append(["💡 Meta realista", "Establece una meta de ahorro del 20% de tus ingresos."])
        elif balance == 0:
            consejos.append(["⚠️ Balance en cero", "Tus ingresos igualan a tus gastos."])
            consejos.append(["💡 Gastos hormiga", "Revisa suscripciones que no uses y compras impulsivas."])
        else:
            consejos.append(["🚨 Alerta financiera", f"Tus gastos superan tus ingresos por {user_currency} {abs(balance):,.2f}."])
            consejos.append(["💡 Plan de acción", "Prioriza gastos esenciales y elimina gastos innecesarios."])
        
        if gastos_categoria:
            mayor_gasto = gastos_categoria[0]
            total_gastos_cat = sum(g['total'] for g in gastos_categoria)
            consejos.append([f"📊 Tu mayor gasto es {mayor_gasto['categoria']}", f"Representa el {(mayor_gasto['total'] / total_gastos_cat * 100):.1f}% de tus gastos."])
        
        presupuestos_excedidos = [p for p in presupuestos if p['gastado'] > p['limite']]
        if presupuestos_excedidos:
            consejos.append(["⚠️ Presupuestos excedidos", f"Has excedido el presupuesto en {len(presupuestos_excedidos)} categoría(s)."])
        
        metas_cumplidas = [m for m in metas if m['porcentaje'] >= 100]
        if metas_cumplidas:
            consejos.append(["🎉 Metas cumplidas", f"¡Felicidades! Has cumplido {len(metas_cumplidas)} meta(s) de ahorro."])
        
        for row, (titulo, consejo) in enumerate(consejos, 3):
            ws_consejos.cell(row=row, column=1, value=titulo)
            ws_consejos.cell(row=row, column=2, value=consejo)
            ws_consejos.cell(row=row, column=1).font = Font(bold=True)
            for col in range(1, 3):
                ws_consejos.cell(row=row, column=col).border = border
                ws_consejos.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical='top')
        
        ws_consejos.column_dimensions['A'].width = 25
        ws_consejos.column_dimensions['B'].width = 50
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=finantrack_reporte_{datetime.now().strftime("%Y%m%d")}.xlsx'
        
        return response
        
    except Exception as e:
        print(f"Error generando Excel: {e}")
        import traceback
        traceback.print_exc()
        if conn:
            conn.close()
        flash('Error al generar el reporte Excel', 'danger')
        return redirect(url_for('dashboard'))

# ==================== RUTAS DE PRUEBA ====================

@app.route('/ver-usuarios')
def ver_usuarios():
    conn = get_db_connection()
    if not conn:
        return "Error de conexión"
    
    try:
        cursor = get_dict_cursor(conn)
        cursor.execute("SELECT id, nombre, email, default_currency, fecha_registro FROM users")
        usuarios = cursor.fetchall()
        cursor.close()
        conn.close()
        
        if not usuarios:
            return "<h3>No hay usuarios</h3><a href='/register'>Registro</a>"
        
        html = "<h2>Usuarios:</h2><ul>"
        for u in usuarios:
            html += f"<li>{u['nombre']} - {u['email']} - Moneda: {u['default_currency']}</li>"
        html += "</ul><a href='/register'>Volver</a>"
        return html
    except Exception as e:
        print(f"Error en ver-usuarios: {e}")
        if conn:
            conn.close()
        return f"Error: {e}"

@app.route('/test-db')
def test_db():
    conn = get_db_connection()
    if not conn:
        return "Error de conexion a la base de datos"
    
    try:
        cursor = conn.cursor()
        if os.environ.get('RENDER'):
            cursor.execute("SELECT version()")
            version = cursor.fetchone()[0]
            cursor.execute("SELECT COUNT(*) FROM users")
            count = cursor.fetchone()[0]
        else:
            cursor.execute("SELECT VERSION()")
            version = cursor.fetchone()[0]
            cursor.execute("SELECT COUNT(*) FROM users")
            count = cursor.fetchone()[0]
        cursor.close()
        conn.close()
        return f"DB OK - Version: {version} - Usuarios: {count}"
    except Exception as e:
        print(f"Error en test-db: {e}")
        if conn:
            conn.close()
        return f"Error: {e}"

# ==================== INICIALIZAR BASE DE DATOS ====================

from init_db_app import initialize_database

# Inicializar la base de datos según el entorno
initialize_database()

# ==================== INICIAR APLICACIÓN ====================

if __name__ == '__main__':
    print("\n" + "="*60)
    print("FINANTRACK - Control de Gastos Personales")
    print("="*60)
    print(f"🌍 Entorno: {'Render' if os.environ.get('RENDER') else 'Desarrollo Local'}")
    print("Servidor: http://127.0.0.1:5000")
    print("Registro: http://127.0.0.1:5000/register")
    print("Divisas: http://127.0.0.1:5000/configuracion-divisas")
    print("="*60 + "\n")
    app.run(debug=True, host='127.0.0.1', port=5000)